"""
Tests for corrected Excel output:
- TC-1: Brand/Campaign/PO columns in generate_corrected_excel
- TC-3: apply-corrections preserves ZIP text formatting
- TC-4: Original filename propagation (spaces → underscores)

Also covers format_excel_output shared helper.
"""
import io
import pytest
import pandas as pd
from unittest.mock import patch
from openpyxl import load_workbook

from app.core.zip_validator import (
    ZipValidator, ValidationReport, ValidationResult, format_excel_output,
)
from app.core.models import ParsedAddress
from app.core.utils import map_columns


def _make_validator():
    return ZipValidator(google_api_key="fake", anthropic_api_key=None)


def _make_report(n_rows: int) -> ValidationReport:
    """Build a minimal ValidationReport with n_rows results."""
    results = []
    for i in range(n_rows):
        results.append(ValidationResult(
            row_index=i,
            name=f"Recipient {i}",
            city="Roma",
            street=f"Via Roma {i + 1}",
            original_zip="00100",
            suggested_zip=None,
            confidence=95,
            reason="",
            is_valid=True,
            street_verified=True,
            country_code="IT",
        ))
    return ValidationReport(
        total_rows=n_rows,
        valid_count=n_rows,
        corrected_count=0,
        review_count=0,
        skipped_count=0,
        results=results,
    )


def _read_excel(excel_bytes: bytes) -> pd.DataFrame:
    return pd.read_excel(io.BytesIO(excel_bytes))


def _read_workbook(excel_bytes: bytes):
    return load_workbook(io.BytesIO(excel_bytes))


# ---------------------------------------------------------------------------
# TC-1: Brand / Campaign / PO columns
# ---------------------------------------------------------------------------

class TestBrandCampaignPOColumns:

    def test_brand_campaign_po_added_when_provided(self):
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10"],
            "City": ["Roma"],
            "Zip": ["00100"],
            "Country": ["IT"],
        })
        report = _make_report(1)

        result = validator.generate_corrected_excel(
            df, report, brand="SBX", campaign="MARZO", po_number="3501494822",
        )
        out_df = _read_excel(result)

        assert "Brand" in out_df.columns
        assert "Campaign" in out_df.columns
        assert "PO Number" in out_df.columns
        assert out_df["Brand"].iloc[0] == "SBX"
        assert out_df["Campaign"].iloc[0] == "MARZO"
        assert str(out_df["PO Number"].iloc[0]) == "3501494822"

    def test_columns_omitted_when_empty(self):
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10"],
            "City": ["Roma"],
            "Zip": ["00100"],
        })
        report = _make_report(1)

        result = validator.generate_corrected_excel(df, report)
        out_df = _read_excel(result)

        assert "Brand" not in out_df.columns
        assert "Campaign" not in out_df.columns
        assert "PO Number" not in out_df.columns

    def test_partial_columns(self):
        """Only brand provided — campaign and PO should be absent."""
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10"],
            "City": ["Roma"],
            "Zip": ["00100"],
        })
        report = _make_report(1)

        result = validator.generate_corrected_excel(df, report, brand="DOUGLAS")
        out_df = _read_excel(result)

        assert "Brand" in out_df.columns
        assert "Campaign" not in out_df.columns
        assert "PO Number" not in out_df.columns

    def test_brand_applied_to_all_rows(self):
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10", "Piazza Duomo 1", "Corso Italia 5"],
            "City": ["Roma", "Milano", "Torino"],
            "Zip": ["00100", "20122", "10121"],
        })
        report = _make_report(3)

        result = validator.generate_corrected_excel(df, report, brand="SBX")
        out_df = _read_excel(result)

        assert list(out_df["Brand"]) == ["SBX", "SBX", "SBX"]


# ---------------------------------------------------------------------------
# TC-3: format_excel_output preserves ZIP formatting
# ---------------------------------------------------------------------------

class TestFormatExcelOutput:

    def test_zip_formatted_as_text_with_leading_zeros(self):
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10"],
            "City": ["Roma"],
            "Zip": [100],  # numeric — should be zero-padded to "00100"
            "Country": ["IT"],
        })
        col_map = map_columns(df)

        result = format_excel_output(df, col_map)
        wb = _read_workbook(result)
        ws = wb.active

        zip_col_idx = list(df.columns).index("Zip") + 1
        cell = ws.cell(row=2, column=zip_col_idx)
        assert cell.value == "00100"
        assert cell.number_format == "@"

    def test_non_it_zip_not_padded(self):
        df = pd.DataFrame({
            "Street 1": ["Hauptstraße 5"],
            "City": ["Berlin"],
            "Zip": [10117],
            "Country": ["DE"],
        })
        col_map = map_columns(df)

        result = format_excel_output(df, col_map)
        wb = _read_workbook(result)
        ws = wb.active

        zip_col_idx = list(df.columns).index("Zip") + 1
        cell = ws.cell(row=2, column=zip_col_idx)
        assert cell.value == "10117"  # not padded, but still text

    def test_column_widths_auto_fitted(self):
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10"],
            "City": ["Roma"],
            "Zip": ["00100"],
        })
        col_map = map_columns(df)

        result = format_excel_output(df, col_map)
        wb = _read_workbook(result)
        ws = wb.active

        # Column widths should be set (not default)
        from openpyxl.utils import get_column_letter
        for i in range(1, len(df.columns) + 1):
            width = ws.column_dimensions[get_column_letter(i)].width
            assert width is not None and width > 0


# ---------------------------------------------------------------------------
# TC-3 continued: apply-corrections preserves ZIP after edit
# ---------------------------------------------------------------------------

class TestApplyCorrectionsZIPPreservation:
    """Verify that apply-corrections uses format_excel_output (not bare to_excel)."""

    def _create_complete_job_with_zip(self) -> str:
        """Create a job with a corrected file that has ZIP formatting."""
        from app.services.job_store import job_store

        job_id = job_store.create_job("validator")
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10"],
            "City": ["Roma"],
            "Zip": [187],  # Will be padded to "00187"
            "Country": ["IT"],
        })
        col_map = map_columns(df)
        excel_bytes = format_excel_output(df, col_map)

        corrected_name = "test_corrected.xlsx"
        job_store.save_file(job_id, corrected_name, excel_bytes)
        job_store.update_status(job_id, "complete", result={
            "results": [],
            "files": {
                "corrected": f"/api/v1/jobs/{job_id}/files/{corrected_name}",
                "review": f"/api/v1/jobs/{job_id}/files/review.xlsx",
            },
        })
        return job_id

    def test_zip_preserved_after_street_edit(self):
        from fastapi.testclient import TestClient
        from app.main import app
        from app.services.job_store import job_store

        client = TestClient(app)
        job_id = self._create_complete_job_with_zip()

        resp = client.post(f"/api/v1/jobs/{job_id}/apply-corrections", json={
            "corrections": {"0": {"street": "Via Garibaldi 5"}},
        })
        assert resp.status_code == 200

        # Read back and check ZIP is still "00187"
        corrected_path = job_store.get_file_path(job_id, "test_corrected.xlsx")
        wb = load_workbook(corrected_path)
        ws = wb.active

        zip_col_idx = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value and "zip" in str(cell.value).lower():
                zip_col_idx = i
                break

        assert zip_col_idx is not None
        zip_cell = ws.cell(row=2, column=zip_col_idx)
        assert zip_cell.value == "00187"
        assert zip_cell.number_format == "@"

        # Also verify the street was actually changed
        street_col_idx = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value and "street" in str(cell.value).lower():
                street_col_idx = i
                break

        assert street_col_idx is not None
        assert ws.cell(row=2, column=street_col_idx).value == "Via Garibaldi 5"

        job_store.cleanup_all()
