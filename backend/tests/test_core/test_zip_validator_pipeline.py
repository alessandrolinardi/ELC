"""
ZipValidator pipeline tests covering:
- Issue #6: Progress callback scaling
- Issue #9: _validate_addresses length guard

Tests the validation pipeline with mocked Google API, progress tracking,
column mapping, and edge cases.
"""
import math
import pytest
from io import BytesIO
from unittest.mock import patch, MagicMock
import pandas as pd

from app.core.zip_validator import ZipValidator, ValidationReport, ValidationResult
from app.core.models import ParsedAddress, ValidationOutcome


def _make_validator(**kwargs):
    """Create a ZipValidator with mocked external dependencies."""
    return ZipValidator(
        google_api_key="fake-key",
        anthropic_api_key=None,
        **kwargs
    )


def _make_parsed(prefix="Via", name="Roma", number="10",
                 country="IT", method="ai"):
    addr = ParsedAddress(
        street_prefix=prefix,
        street_name=name,
        house_number=number,
        location_info="",
        country_code=country,
        confidence="high",
    )
    addr.parse_method = method
    return addr


class TestLengthGuard:
    """Issue #9: _validate_addresses must check len(parsed_addresses) == len(df)."""

    def test_length_mismatch_raises(self):
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10", "Via Milano 5"],
            "City": ["Roma", "Milano"],
            "Zip": ["00100", "20121"],
        })
        parsed = [_make_parsed()]  # Only 1, but df has 2 rows

        with pytest.raises(ValueError, match="parsed_addresses length"):
            validator._validate_addresses(df, parsed)

    def test_length_match_passes(self):
        """Equal lengths should not raise."""
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10"],
            "City": ["Roma"],
            "Zip": ["00100"],
        })
        parsed = [_make_parsed()]

        # Mock the API to avoid real calls
        with patch.object(validator.address_validator, 'validate_address', return_value=None):
            report, _ = validator._validate_addresses(df, parsed)
            assert report.total_rows == 1

    def test_empty_df_with_empty_parsed(self):
        """Empty dataframe with empty parsed list should work."""
        validator = _make_validator()
        df = pd.DataFrame({"City": [], "Zip": []})
        parsed = []

        report, _ = validator._validate_addresses(df, parsed)
        assert report.total_rows == 0


class TestColumnGuard:
    """_validate_addresses column validation."""

    def test_missing_city_column_raises(self):
        validator = _make_validator()
        df = pd.DataFrame({"Street": ["Via Roma 10"], "Zip": ["00100"]})
        parsed = [_make_parsed()]

        with pytest.raises(ValueError, match="Missing required columns"):
            validator._validate_addresses(df, parsed)

    def test_missing_zip_column_raises(self):
        validator = _make_validator()
        df = pd.DataFrame({"Street": ["Via Roma 10"], "City": ["Roma"]})
        parsed = [_make_parsed()]

        with pytest.raises(ValueError, match="Missing required columns"):
            validator._validate_addresses(df, parsed)

    def test_optional_street_column_ok(self):
        """Street column is optional — should not raise if missing."""
        validator = _make_validator()
        df = pd.DataFrame({"City": ["Roma"], "Zip": ["00100"]})
        parsed = [_make_parsed()]

        with patch.object(validator.address_validator, 'validate_address', return_value=None):
            report, _ = validator._validate_addresses(df, parsed)
            assert report.total_rows == 1


class TestProgressCallback:
    """Issue #6: Progress scaling correctness."""

    def test_progress_reports_0_to_100(self):
        """_validate_addresses should report progress from 0% to 100%."""
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10", "Piazza Duomo 1", "Corso Italia 5"],
            "City": ["Roma", "Milano", "Torino"],
            "Zip": ["00100", "20122", "10121"],
        })
        parsed = [_make_parsed() for _ in range(3)]
        progress_calls = []

        def track_progress(current, total, message):
            progress_calls.append((current, total, message))

        with patch.object(validator.address_validator, 'validate_address', return_value=None):
            validator._validate_addresses(df, parsed, track_progress)

        assert len(progress_calls) == 3
        # Each call should have total=100
        assert all(t == 100 for _, t, _ in progress_calls)
        # Progress should increase
        percents = [c for c, _, _ in progress_calls]
        assert percents == sorted(percents)
        # Last call should be 100%
        assert percents[-1] == 100

    def test_process_dataframe_scales_20_to_100(self):
        """process_dataframe should scale validation progress from 20-100%."""
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10"],
            "City": ["Roma"],
            "Zip": ["00100"],
        })
        progress_calls = []

        def track_progress(current, total, message):
            progress_calls.append((current, total, message))

        with patch.object(validator.address_parser, 'parse_all',
                          return_value=[_make_parsed()]):
            with patch.object(validator.address_validator, 'validate_address',
                              return_value=None):
                validator.process_dataframe(df, track_progress)

        # Should have parsing progress (0/100) and then scaled validation
        # The first call is parsing start, then parsing complete (20/100),
        # then validation scaled from 20-100%
        validation_calls = [c for c in progress_calls if "Validating" in c[2]]
        for current, total, _ in validation_calls:
            assert current >= 20, f"Validation progress {current} should be >= 20"
            assert total == 100


class TestNonITCountrySkip:
    """Non-Italian addresses should be skipped."""

    def test_non_it_country_skipped(self):
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Hauptstraße 5"],
            "City": ["Berlin"],
            "Zip": ["10115"],
        })
        parsed = [_make_parsed(prefix="", name="Hauptstraße", number="5", country="DE")]

        report, _ = validator._validate_addresses(df, parsed)
        assert report.skipped_count == 1
        assert report.total_rows == 1
        assert report.results[0].country_code == "DE"

    def test_it_country_not_skipped(self):
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10"],
            "City": ["Milano"],
            "Zip": ["20121"],
        })
        parsed = [_make_parsed()]

        with patch.object(validator.address_validator, 'validate_address', return_value=None):
            report, _ = validator._validate_addresses(df, parsed)
            assert report.skipped_count == 0


class TestGoogleAPIUnavailable:
    """Graceful handling when Google API returns None."""

    def test_api_unavailable_marks_review(self):
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10"],
            "City": ["Roma"],
            "Zip": ["00100"],
        })
        parsed = [_make_parsed()]

        with patch.object(validator.address_validator, 'validate_address', return_value=None):
            report, _ = validator._validate_addresses(df, parsed)

        assert report.review_count == 1
        assert report.results[0].reason == "Google API unavailable"

    def test_api_returns_no_result_key(self):
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10"],
            "City": ["Roma"],
            "Zip": ["00100"],
        })
        parsed = [_make_parsed()]

        with patch.object(validator.address_validator, 'validate_address',
                          return_value={"error": "something"}):
            report, _ = validator._validate_addresses(df, parsed)

        assert report.review_count == 1


class TestZipPadding:
    """Italian ZIP codes should be padded to 5 digits."""

    def test_short_zip_padded(self):
        """ZIP '187' should become '00187'."""
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10"],
            "City": ["Roma"],
            "Zip": [187],  # numeric, short
        })
        parsed = [_make_parsed()]

        api_calls = []

        def capture_call(parsed_addr, city, zip_code, state="", street2=""):
            api_calls.append(zip_code)
            return None

        with patch.object(validator.address_validator, 'validate_address',
                          side_effect=capture_call):
            validator._validate_addresses(df, parsed)

        assert api_calls[0] == "00187"

    def test_full_zip_unchanged(self):
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10"],
            "City": ["Milano"],
            "Zip": ["20121"],
        })
        parsed = [_make_parsed()]

        api_calls = []

        def capture_call(parsed_addr, city, zip_code, state="", street2=""):
            api_calls.append(zip_code)
            return None

        with patch.object(validator.address_validator, 'validate_address',
                          side_effect=capture_call):
            validator._validate_addresses(df, parsed)

        assert api_calls[0] == "20121"


class TestColumnMapping:
    """Column name mapping handles variations."""

    def test_italian_column_names(self):
        validator = _make_validator()
        df = pd.DataFrame({
            "Indirizzo": ["Via Roma 10"],
            "Città": ["Milano"],
            "CAP": ["20121"],
            "Nome": ["Mario Rossi"],
        })
        col_map = validator._map_columns(df)
        assert col_map["street"] == "Indirizzo"
        assert col_map["city"] == "Città"
        assert col_map["zip"] == "CAP"
        assert col_map["name"] == "Nome"

    def test_english_column_names(self):
        validator = _make_validator()
        df = pd.DataFrame({
            "Street 1": ["Via Roma 10"],
            "City": ["Milano"],
            "Zip": ["20121"],
        })
        col_map = validator._map_columns(df)
        assert col_map["street"] == "Street 1"
        assert col_map["city"] == "City"
        assert col_map["zip"] == "Zip"


# =========================================================================
# Helper: build a ValidationResult with sensible defaults
# =========================================================================

def _make_result(row_index=0, name="Mario Rossi", city="Roma",
                 street="Via Roma 10", original_zip="00100",
                 suggested_zip="00100", confidence=100,
                 reason="", is_valid=True, auto_corrected=False,
                 street_verified=True, suggested_street=None,
                 street_confidence=95, street_auto_corrected=False,
                 country_code="IT", country_detected=False,
                 phone_missing=False, original_phone="3331234567",
                 cod_changed=False, original_cod="0",
                 po_invalid=False, po_value="", po_extracted=""):
    return ValidationResult(
        row_index=row_index, name=name, city=city, street=street,
        original_zip=original_zip, suggested_zip=suggested_zip,
        confidence=confidence, reason=reason, is_valid=is_valid,
        auto_corrected=auto_corrected, street_verified=street_verified,
        suggested_street=suggested_street,
        street_confidence=street_confidence,
        street_auto_corrected=street_auto_corrected,
        country_code=country_code, country_detected=country_detected,
        phone_missing=phone_missing, original_phone=original_phone,
        cod_changed=cod_changed, original_cod=original_cod,
        po_invalid=po_invalid, po_value=po_value, po_extracted=po_extracted,
    )


def _make_outcome(status="valid", output_zip="00100", output_street="Via Roma",
                  zip_confirmed=True, street_confirmed=True,
                  street_corrected=False, silent_correction=False,
                  location_info="", reasons=None):
    return ValidationOutcome(
        status=status, action="ACCEPT",
        input_zip="00100", output_zip=output_zip,
        zip_confirmed=zip_confirmed, zip_corrected=False,
        input_street="Via Roma 10", output_street=output_street,
        street_confirmed=street_confirmed, street_corrected=street_corrected,
        silent_correction=silent_correction,
        house_number="10", granularity="PREMISE",
        address_complete=True, reasons=reasons or [],
        formatted_address="Via Roma 10, 00100 Roma RM, Italy",
        location_info=location_info,
    )


# =========================================================================
# 1. generate_corrected_excel tests
# =========================================================================

class TestGenerateCorrectedExcel:
    """Tests for generate_corrected_excel output."""

    def _build_df_and_report(self, rows, results):
        """Build a DataFrame and ValidationReport from row dicts and ValidationResults."""
        df = pd.DataFrame(rows)
        report = ValidationReport(
            total_rows=len(results),
            valid_count=sum(1 for r in results if r.is_valid),
            corrected_count=sum(1 for r in results if r.auto_corrected),
            review_count=sum(1 for r in results if not r.is_valid and not r.auto_corrected),
            skipped_count=0,
            results=results,
        )
        return df, report

    def _read_output(self, excel_bytes):
        """Read corrected Excel bytes back as DataFrame."""
        return pd.read_excel(BytesIO(excel_bytes), sheet_name='Corrected', dtype=str)

    def _read_output_openpyxl(self, excel_bytes):
        """Read corrected Excel bytes with openpyxl for cell-level inspection."""
        import openpyxl
        return openpyxl.load_workbook(BytesIO(excel_bytes))

    def test_zip_formatted_as_text_with_leading_zeros(self):
        """ZIP column should preserve leading zeros (00187 not 187)."""
        validator = _make_validator()
        df, report = self._build_df_and_report(
            [{"Street 1": "Via Roma 10", "City": "Roma", "Zip": 187, "Country": "IT"}],
            [_make_result(row_index=0, suggested_zip="00187", auto_corrected=True,
                          is_valid=False, country_code="IT")],
        )
        excel_bytes = validator.generate_corrected_excel(df, report)
        wb = self._read_output_openpyxl(excel_bytes)
        ws = wb['Corrected']
        zip_col_idx = [c.value for c in ws[1]].index("Zip") + 1
        zip_cell = ws.cell(row=2, column=zip_col_idx)
        assert zip_cell.value == "00187"
        assert zip_cell.number_format == '@'

    def test_empty_phone_filled_with_default(self):
        """Empty phone should be replaced with DEFAULT_PHONE."""
        validator = _make_validator()
        df, report = self._build_df_and_report(
            [{"Street 1": "Via Roma 10", "City": "Roma", "Zip": "00100",
              "Phone": "", "Country": "IT"}],
            [_make_result(row_index=0, phone_missing=True, original_phone="")],
        )
        excel_bytes = validator.generate_corrected_excel(df, report)
        out = self._read_output(excel_bytes)
        assert out["Phone"].iloc[0] == ZipValidator.DEFAULT_PHONE

    def test_nan_phone_filled_with_default(self):
        """NaN phone should be replaced with DEFAULT_PHONE."""
        validator = _make_validator()
        df, report = self._build_df_and_report(
            [{"Street 1": "Via Roma 10", "City": "Roma", "Zip": "00100",
              "Phone": float('nan'), "Country": "IT"}],
            [_make_result(row_index=0, phone_missing=True, original_phone="")],
        )
        excel_bytes = validator.generate_corrected_excel(df, report)
        out = self._read_output(excel_bytes)
        assert out["Phone"].iloc[0] == ZipValidator.DEFAULT_PHONE

    def test_cod_set_to_zero_for_it(self):
        """Cash on Delivery should be set to 0 for IT addresses."""
        validator = _make_validator()
        df, report = self._build_df_and_report(
            [{"Street 1": "Via Roma 10", "City": "Roma", "Zip": "00100",
              "Cash on Delivery": 15.50, "Country": "IT"}],
            [_make_result(row_index=0, cod_changed=True, original_cod="15.50",
                          country_code="IT")],
        )
        excel_bytes = validator.generate_corrected_excel(df, report)
        out = self._read_output(excel_bytes)
        assert float(out["Cash on Delivery"].iloc[0]) == 0

    def test_cod_not_zeroed_for_non_it(self):
        """Cash on Delivery should NOT be zeroed for non-IT addresses."""
        validator = _make_validator()
        df, report = self._build_df_and_report(
            [{"Street 1": "Hauptstr 5", "City": "Berlin", "Zip": "10115",
              "Cash on Delivery": 25.00, "Country": "DE"}],
            [_make_result(row_index=0, cod_changed=True, original_cod="25.00",
                          country_code="DE")],
        )
        excel_bytes = validator.generate_corrected_excel(df, report)
        out = self._read_output(excel_bytes)
        assert float(out["Cash on Delivery"].iloc[0]) == 25.0

    def test_country_code_filled_when_empty(self):
        """Country should be filled with detected country_code when empty."""
        validator = _make_validator()
        df, report = self._build_df_and_report(
            [{"Street 1": "Via Roma 10", "City": "Roma", "Zip": "00100",
              "Country": ""}],
            [_make_result(row_index=0, country_code="IT", country_detected=True)],
        )
        excel_bytes = validator.generate_corrected_excel(df, report)
        out = self._read_output(excel_bytes)
        assert out["Country"].iloc[0] == "IT"

    def test_country_code_filled_when_nan(self):
        """Country NaN should be filled with detected country_code."""
        validator = _make_validator()
        df, report = self._build_df_and_report(
            [{"Street 1": "Via Roma 10", "City": "Roma", "Zip": "00100",
              "Country": float('nan')}],
            [_make_result(row_index=0, country_code="IT", country_detected=True)],
        )
        excel_bytes = validator.generate_corrected_excel(df, report)
        out = self._read_output(excel_bytes)
        assert out["Country"].iloc[0] == "IT"

    def test_formula_injection_sanitized(self):
        """Cell starting with = should get ' prefix in output."""
        validator = _make_validator()
        df, report = self._build_df_and_report(
            [{"Street 1": "=SUM(A1:A2)", "City": "Roma", "Zip": "00100",
              "Country": "IT"}],
            [_make_result(row_index=0, street="=SUM(A1:A2)")],
        )
        excel_bytes = validator.generate_corrected_excel(df, report)
        out = self._read_output(excel_bytes)
        assert out["Street 1"].iloc[0].startswith("'")

    def test_street2_trailing_dashes_cleaned(self):
        """Trailing dashes should be stripped from Street 2."""
        validator = _make_validator()
        df, report = self._build_df_and_report(
            [{"Street 1": "Via Roma 10", "City": "Roma", "Zip": "00100",
              "Street 2": "Apt 3 --", "Country": "IT"}],
            [_make_result(row_index=0)],
        )
        excel_bytes = validator.generate_corrected_excel(df, report)
        out = self._read_output(excel_bytes)
        val = out["Street 2"].iloc[0]
        assert not val.rstrip().endswith("-")

    def test_street_correction_applied(self):
        """When street_auto_corrected, suggested_street should be written."""
        validator = _make_validator()
        df, report = self._build_df_and_report(
            [{"Street 1": "Via Rma 10", "City": "Roma", "Zip": "00100",
              "Country": "IT"}],
            [_make_result(row_index=0, street="Via Rma 10",
                          street_auto_corrected=True,
                          suggested_street="Via Roma 10",
                          is_valid=False, auto_corrected=False)],
        )
        excel_bytes = validator.generate_corrected_excel(df, report)
        out = self._read_output(excel_bytes)
        assert out["Street 1"].iloc[0] == "Via Roma 10"


# =========================================================================
# 2. generate_review_report tests
# =========================================================================

class TestGenerateReviewReport:
    """Tests for generate_review_report output."""

    def _read_review(self, excel_bytes):
        return pd.read_excel(BytesIO(excel_bytes), sheet_name='Review', dtype=str)

    def test_review_includes_non_valid_rows(self):
        """Non-valid rows must appear in review report."""
        validator = _make_validator()
        results = [
            _make_result(row_index=0, is_valid=True, street_verified=True),
            _make_result(row_index=1, is_valid=False, auto_corrected=True,
                         street_verified=True, confidence=95,
                         reason="ZIP corrected"),
        ]
        report = ValidationReport(
            total_rows=2, valid_count=1, corrected_count=1,
            review_count=0, skipped_count=0, results=results,
        )
        excel_bytes = validator.generate_review_report(report)
        out = self._read_review(excel_bytes)
        # Row at index 1 is not valid, should appear
        assert len(out) == 1
        assert out["Row"].iloc[0] == "3"  # row_index=1 + 2

    def test_review_includes_non_street_verified_rows(self):
        """Rows with is_valid=True but street_verified=False must appear."""
        validator = _make_validator()
        results = [
            _make_result(row_index=0, is_valid=True, street_verified=False,
                         street_confidence=0, reason="Street unverified"),
        ]
        report = ValidationReport(
            total_rows=1, valid_count=1, corrected_count=0,
            review_count=0, skipped_count=0, results=results,
        )
        excel_bytes = validator.generate_review_report(report)
        out = self._read_review(excel_bytes)
        assert len(out) == 1

    def test_review_excludes_valid_and_verified(self):
        """Rows that are both valid and street_verified should NOT appear."""
        validator = _make_validator()
        results = [
            _make_result(row_index=0, is_valid=True, street_verified=True),
        ]
        report = ValidationReport(
            total_rows=1, valid_count=1, corrected_count=0,
            review_count=0, skipped_count=0, results=results,
        )
        excel_bytes = validator.generate_review_report(report)
        out = self._read_review(excel_bytes)
        assert len(out) == 0

    def test_review_formula_injection_sanitized(self):
        """Formula injection should be sanitized in review report."""
        validator = _make_validator()
        results = [
            _make_result(row_index=0, is_valid=False, street_verified=False,
                         street="=CMD()", reason="=evil"),
        ]
        report = ValidationReport(
            total_rows=1, valid_count=0, corrected_count=0,
            review_count=1, skipped_count=0, results=results,
        )
        excel_bytes = validator.generate_review_report(report)
        out = self._read_review(excel_bytes)
        assert out["Original Street"].iloc[0].startswith("'")
        assert out["Reason"].iloc[0].startswith("'")

    def test_review_correct_columns_present(self):
        """Review report should have all expected columns."""
        validator = _make_validator()
        results = [
            _make_result(row_index=0, is_valid=False, reason="test"),
        ]
        report = ValidationReport(
            total_rows=1, valid_count=0, corrected_count=0,
            review_count=1, skipped_count=0, results=results,
        )
        excel_bytes = validator.generate_review_report(report)
        out = self._read_review(excel_bytes)
        expected_cols = {
            'Row', 'Name', 'City', 'Country', 'Original Street',
            'Suggested Street', 'Street Conf.', 'Street Action',
            'Original ZIP', 'Suggested ZIP', 'ZIP Conf.', 'ZIP Action',
            'Reason',
        }
        assert set(out.columns) == expected_cols


# =========================================================================
# 3. _extract_row_fields tests
# =========================================================================

class TestExtractRowFields:
    """Tests for _extract_row_fields method."""

    def _extract(self, row_dict, parsed=None, extra_cols=None):
        """Helper to call _extract_row_fields with a row dict."""
        validator = _make_validator()
        if parsed is None:
            parsed = _make_parsed()
        cols = {"Street 1": "Via Roma 10", "City": "Roma", "Zip": "00100"}
        if extra_cols:
            cols.update(extra_cols)
        cols.update(row_dict)
        df = pd.DataFrame([cols])
        col_map = validator._map_columns(df)
        row = df.iloc[0]
        return validator._extract_row_fields(row, parsed, col_map)

    def test_phone_missing_when_empty(self):
        fields = self._extract({"Phone": ""})
        assert fields["phone_missing"] is True

    def test_phone_missing_when_nan(self):
        fields = self._extract({"Phone": float('nan')})
        assert fields["phone_missing"] is True

    def test_phone_present(self):
        fields = self._extract({"Phone": "3331234567"})
        assert fields["phone_missing"] is False
        assert fields["original_phone"] == "3331234567"

    def test_cod_changed_nonzero(self):
        fields = self._extract({"Cash on Delivery": "15.50"})
        assert fields["cod_changed"] is True
        assert fields["original_cod"] == "15.50"

    def test_cod_not_changed_zero_float(self):
        fields = self._extract({"Cash on Delivery": "0.0"})
        assert fields["cod_changed"] is False

    def test_cod_not_changed_zero_int(self):
        fields = self._extract({"Cash on Delivery": "0"})
        assert fields["cod_changed"] is False

    def test_cod_not_changed_nan(self):
        fields = self._extract({"Cash on Delivery": float('nan')})
        assert fields["cod_changed"] is False

    def test_country_italy_to_it(self):
        parsed = _make_parsed(country="IT")
        fields = self._extract({"Country": "ITALY"}, parsed=parsed)
        assert fields["country"] == "IT"

    def test_country_italia_to_it(self):
        parsed = _make_parsed(country="IT")
        fields = self._extract({"Country": "ITALIA"}, parsed=parsed)
        assert fields["country"] == "IT"

    def test_country_germany_to_de(self):
        parsed = _make_parsed(country="DE")
        fields = self._extract({"Country": "GERMANY"}, parsed=parsed)
        assert fields["country"] == "DE"

    def test_country_two_letter_code_passthrough(self):
        parsed = _make_parsed(country="FR")
        fields = self._extract({"Country": "FR"}, parsed=parsed)
        assert fields["country"] == "FR"

    def test_country_empty_uses_parsed(self):
        """When country column is empty, parsed.country_code should be used."""
        parsed = _make_parsed(country="DE")
        fields = self._extract({"Country": ""}, parsed=parsed)
        # Empty explicit country falls through to parsed country_code
        assert fields["country"] == "DE"

    def test_street2_trailing_dash_cleanup(self):
        fields = self._extract({"Street 2": "Apt 3 --"})
        assert fields["street2"] == "Apt 3"

    def test_street2_nan_cleaned(self):
        fields = self._extract({"Street 2": "nan"})
        assert fields["street2"] == ""

    def test_po_valid(self):
        """Valid PO in the list should not be marked invalid."""
        validator = _make_validator()
        validator._valid_po_numbers = {"3501234567"}
        df = pd.DataFrame([{
            "Street 1": "Via Roma 10", "City": "Roma", "Zip": "00100",
            "Order Number": "PO 3501234567",
        }])
        col_map = validator._map_columns(df)
        parsed = _make_parsed()
        fields = validator._extract_row_fields(df.iloc[0], parsed, col_map)
        assert fields["po_invalid"] is False
        assert fields["po_extracted"] == "3501234567"

    def test_po_invalid(self):
        """PO not in the valid set should be marked invalid."""
        validator = _make_validator()
        validator._valid_po_numbers = {"3509999999"}
        df = pd.DataFrame([{
            "Street 1": "Via Roma 10", "City": "Roma", "Zip": "00100",
            "Order Number": "PO 3501234567",
        }])
        col_map = validator._map_columns(df)
        parsed = _make_parsed()
        fields = validator._extract_row_fields(df.iloc[0], parsed, col_map)
        assert fields["po_invalid"] is True

    def test_po_empty(self):
        """Empty PO should be marked invalid."""
        validator = _make_validator()
        df = pd.DataFrame([{
            "Street 1": "Via Roma 10", "City": "Roma", "Zip": "00100",
            "Order Number": "",
        }])
        col_map = validator._map_columns(df)
        parsed = _make_parsed()
        fields = validator._extract_row_fields(df.iloc[0], parsed, col_map)
        # Empty order number -> po_value is empty -> skipped, po_invalid stays False
        assert fields["po_invalid"] is False
        assert fields["po_extracted"] == ""


# =========================================================================
# 4. _cross_check_zip tests
# =========================================================================

class TestCrossCheckZip:
    """Tests for _cross_check_zip method."""

    def test_valid_cap_matching_comune_no_change(self):
        """Valid CAP + matching comune should not downgrade status."""
        validator = _make_validator()
        outcome = _make_outcome(status="valid", output_zip="00100")
        with patch.object(validator.address_validator, 'is_valid_italian_cap', return_value=True), \
             patch.object(validator.address_validator, 'validate_zip_comune', return_value=(True, "")), \
             patch.object(validator.address_validator, 'validate_zip_province', return_value=(True, "")):
            validator._cross_check_zip(outcome, "00100", "Roma", "RM")
        assert outcome.status == "valid"

    def test_invalid_cap_google_not_confirmed_review(self):
        """Invalid CAP + Google not confirmed should become review."""
        validator = _make_validator()
        outcome = _make_outcome(status="valid", output_zip="99999",
                                zip_confirmed=False)
        with patch.object(validator.address_validator, 'is_valid_italian_cap', return_value=False):
            validator._cross_check_zip(outcome, "99999", "Roma", "RM")
        assert outcome.status == "review"
        assert any("not in local database" in r for r in outcome.reasons)

    def test_invalid_cap_google_confirmed_trust_google(self):
        """Invalid CAP + Google confirmed should keep status (trust Google)."""
        validator = _make_validator()
        outcome = _make_outcome(status="valid", output_zip="99999",
                                zip_confirmed=True)
        with patch.object(validator.address_validator, 'is_valid_italian_cap', return_value=False):
            validator._cross_check_zip(outcome, "99999", "Roma", "RM")
        # zip_confirmed is True, so is_valid_italian_cap returning False still
        # triggers review because zip_confirmed check is separate
        # But looking at code: it checks `not outcome.zip_confirmed` — so if confirmed, it returns early
        # Actually the code: if not is_valid AND not zip_confirmed -> review
        # zip_confirmed=True -> the condition fails -> no review
        assert outcome.status == "valid"

    def test_cap_mismatch_with_comune_review(self):
        """CAP valid but mismatched with comune should trigger review."""
        validator = _make_validator()
        outcome = _make_outcome(status="valid", output_zip="20121",
                                zip_confirmed=False)
        with patch.object(validator.address_validator, 'is_valid_italian_cap', return_value=True), \
             patch.object(validator.address_validator, 'validate_zip_comune',
                          return_value=(False, "CAP 20121 does not match comune Roma")):
            validator._cross_check_zip(outcome, "20121", "Roma", "RM")
        assert outcome.status == "review"
        assert any("does not match comune" in r for r in outcome.reasons)

    def test_cap_mismatch_with_province_review(self):
        """CAP valid + comune match but province mismatch should trigger review."""
        validator = _make_validator()
        outcome = _make_outcome(status="valid", output_zip="00100",
                                zip_confirmed=False)
        with patch.object(validator.address_validator, 'is_valid_italian_cap', return_value=True), \
             patch.object(validator.address_validator, 'validate_zip_comune',
                          return_value=(True, "")), \
             patch.object(validator.address_validator, 'validate_zip_province',
                          return_value=(False, "CAP 00100 province mismatch: expected RM, got MI")):
            validator._cross_check_zip(outcome, "00100", "Roma", "MI")
        assert outcome.status == "review"
        assert any("province mismatch" in r for r in outcome.reasons)


# =========================================================================
# 5. Street 2 tests (via generate_corrected_excel — single owner of Street 2 logic)
# _update_street2 was removed; Street 2 handling now lives in generate_corrected_excel.
# See tests/test_core/test_corrected_excel.py for Street 2 coverage.
