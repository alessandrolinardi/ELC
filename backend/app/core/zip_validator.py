"""
Address Validator Module
Validates and corrects addresses using Claude AI for parsing
and Google Address Validation API for validation.

Includes caching layer with Supabase for faster repeated validations.
"""

import re
import json
from dataclasses import dataclass
from typing import Optional, Callable
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta, timezone

import pandas as pd

from .logging_config import get_logger
from .models import ParsedAddress, ValidationOutcome, ParsingMetrics
from .address_parser import AddressParser
from .address_validator import AddressValidator
from .utils import map_columns, sanitize_cell

# Logger for this module
logger = get_logger(__name__)


def format_excel_output(df: pd.DataFrame, col_map: dict) -> bytes:
    """Write DataFrame to Excel with ZIP text formatting and auto-fit columns.

    Shared by generate_corrected_excel and apply-corrections to ensure
    consistent output formatting (ZIP as text with leading zeros, column widths).
    """
    from openpyxl.utils import get_column_letter

    zip_col = col_map.get('zip')
    country_col = col_map.get('country')

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Corrected')
        worksheet = writer.sheets['Corrected']

        # Format ZIP column as text with leading zeros preserved (IT only)
        if zip_col and zip_col in df.columns:
            zip_col_idx = list(df.columns).index(zip_col) + 1
            country_col_idx = list(df.columns).index(country_col) + 1 if country_col and country_col in df.columns else None
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=zip_col_idx)
                cell.number_format = '@'
                raw = cell.value
                if raw is not None:
                    is_it = True
                    if country_col_idx:
                        country_cell = worksheet.cell(row=row, column=country_col_idx)
                        is_it = str(country_cell.value).upper().strip() in ('IT', 'ITALY', 'ITALIA', '')
                    if is_it:
                        try:
                            cell.value = str(int(float(str(raw)))).zfill(5)
                        except (ValueError, TypeError):
                            cell.value = str(raw)
                    else:
                        cell.value = str(raw)

        # Auto-fit column widths
        for col_idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                len(str(col))
            ) + 2
            column_width = min(max_length, 50)
            worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = column_width

    return output.getvalue()


@dataclass
class ValidationResult:
    """Result of a single address validation."""
    row_index: int
    name: str
    city: str
    street: str
    original_zip: str
    suggested_zip: Optional[str]
    confidence: int  # 0-100
    reason: str
    is_valid: bool
    auto_corrected: bool = False
    # Street validation fields
    street_verified: bool = False
    suggested_street: Optional[str] = None
    street_confidence: int = 0
    street_auto_corrected: bool = False
    # Country code
    country_code: str = "IT"
    country_detected: bool = False
    # Additional corrections tracking
    phone_missing: bool = False
    original_phone: str = ""
    cod_changed: bool = False
    original_cod: str = ""
    # Location info (C.C., outlet names, etc.) extracted by AI parser
    location_info: str = ""
    # PO validation
    po_invalid: bool = False
    po_value: str = ""
    po_extracted: str = ""


@dataclass
class ValidationReport:
    """Complete validation report."""
    total_rows: int
    valid_count: int
    corrected_count: int
    review_count: int
    skipped_count: int
    results: list[ValidationResult]
    street_verified_count: int = 0
    street_corrected_count: int = 0
    po_invalid_count: int = 0


class ZipValidator:
    """
    Validates addresses using Claude AI for parsing
    and Google Address Validation API for validation.
    """

    # Default phone number to use when phone is missing
    DEFAULT_PHONE = "393445556667"

    def __init__(self, confidence_threshold: int = 90, street_confidence_threshold: int = 85,
                 google_api_key: Optional[str] = None, anthropic_api_key: Optional[str] = None):
        """
        Initialize validator.

        Args:
            confidence_threshold: Minimum confidence for auto-correction of ZIP (default 90%)
            street_confidence_threshold: Minimum confidence for auto-correction of street (default 85%)
            google_api_key: Google Address Validation API key
            anthropic_api_key: Anthropic API key for Claude address parsing
        """
        self.confidence_threshold = confidence_threshold
        self.street_confidence_threshold = street_confidence_threshold
        self.address_parser = AddressParser(api_key=anthropic_api_key)
        self.address_validator = AddressValidator(api_key=google_api_key)
        self._valid_po_numbers = self._load_valid_po_numbers()
        # Supabase client removed — address cache was write-only (read path
        # was dead code), and cache key excluded ZIP, causing wrong results.

    # =========================================================================
    # Cache methods
    # =========================================================================

    # Address cache removed — was write-only (lookup never called),
    # cache key excluded ZIP (wrong results for same street+city with different ZIP),
    # and added Supabase dependency overhead for zero benefit.

    # =========================================================================
    # PO validation
    # =========================================================================

    def _load_valid_po_numbers(self) -> set:
        """Load valid PO numbers from JSON file."""
        po_file = Path(__file__).parent.parent.parent / "data" / "valid_po_numbers.json"
        try:
            if po_file.exists():
                with open(po_file, 'r') as f:
                    data = json.load(f)
                    return set(data.get('po_numbers', []))
        except (json.JSONDecodeError, IOError) as e:
            logger.warning(f"Could not load PO numbers: {e}")
        return set()

    def extract_po_from_string(self, value: str) -> Optional[str]:
        """Extract PO number from a string. PO numbers are 10-digit numbers starting with 350."""
        if not value:
            return None
        matches = re.findall(r'350\d{7}', str(value))
        if matches:
            return matches[0]
        return None

    def validate_po_number(self, order_number: str) -> tuple[bool, str, str]:
        """Validate that an Order Number contains a valid PO number."""
        if not order_number or str(order_number).lower() in ('nan', 'none', ''):
            return False, "", "Order Number vuoto"

        extracted_po = self.extract_po_from_string(str(order_number))
        if not extracted_po:
            return False, "", f"Nessun PO trovato in '{order_number}'"
        if extracted_po not in self._valid_po_numbers:
            return False, extracted_po, f"PO {extracted_po} non valido (non in lista)"
        return True, extracted_po, ""

    # =========================================================================
    # Regex fallback (kept for Tier 2/3)

    # =========================================================================
    # Column mapping (unchanged)
    # =========================================================================

    def _map_columns(self, df: pd.DataFrame) -> dict:
        """Map DataFrame columns to expected fields."""
        return map_columns(df)

    # =========================================================================
    # Main pipeline
    # =========================================================================

    def process_dataframe(
        self,
        df: pd.DataFrame,
        progress_callback: Optional[Callable[[int, int, str], None]] = None
    ) -> tuple[ValidationReport, pd.DataFrame]:
        """
        Process entire DataFrame and validate all addresses.

        Args:
            df: DataFrame with address data
            progress_callback: Optional callback(current, total, message)

        Returns:
            Tuple of (ValidationReport with all results, DataFrame)
        """
        logger.info(f"Starting address validation for {len(df)} rows")

        col_map = self._map_columns(df)
        logger.debug(f"Column mapping: {col_map}")

        if not all(col_map.get(k) for k in ['city', 'zip']):
            raise ValueError(
                f"Missing required columns. Found: {list(df.columns)}\n"
                f"Need: City, Zip"
            )

        # --- Step 1: Claude batch parse all addresses ---
        if progress_callback:
            progress_callback(0, 100, "Parsing addresses with AI...")

        raw_addresses = []
        for idx, row in df.iterrows():
            street = str(row.get(col_map.get('street', ''), ''))
            city = str(row.get(col_map['city'], ''))
            zip_val = str(row.get(col_map['zip'], ''))
            raw_addresses.append({"street": street, "city": city, "zip": zip_val})

        parsed_addresses = self.address_parser.parse_all(raw_addresses)

        if progress_callback:
            progress_callback(20, 100, "Parsing complete")

        logger.info(f"Parsing complete: {self.address_parser.metrics.claude_parsed} by Claude, "
                     f"{self.address_parser.metrics.regex_fallback} by regex")

        # Step 2: Validate (delegated) — scale 0-100% from _validate_addresses to 20-100%
        def scaled_callback(current, total, message):
            if progress_callback:
                pct = 20 + int(current / total * 80) if total > 0 else 20
                progress_callback(pct, 100, message)

        return self._validate_addresses(df, parsed_addresses, scaled_callback)

    def _validate_addresses(
        self,
        df: pd.DataFrame,
        parsed_addresses: list,
        progress_callback: Optional[Callable[[int, int, str], None]] = None
    ) -> tuple:
        """Run Google validation on pre-parsed addresses. Skips parsing step.

        Args:
            df: Original DataFrame with address data
            parsed_addresses: List of ParsedAddress objects (one per row)
            progress_callback: Optional callback(current, total, message)

        Returns:
            tuple[ValidationReport, pd.DataFrame] — same as process_dataframe()
        """
        col_map = self._map_columns(df)
        required = ['city', 'zip']
        missing = [k for k in required if k not in col_map]
        if missing:
            raise ValueError(f"Missing required columns: {missing}")

        if len(parsed_addresses) != len(df):
            raise ValueError(
                f"parsed_addresses length ({len(parsed_addresses)}) != "
                f"DataFrame rows ({len(df)})"
            )

        results = []
        valid_count = 0
        corrected_count = 0
        review_count = 0
        skipped_count = 0
        street_verified_count = 0
        street_corrected_count = 0
        po_invalid_count = 0
        total = len(df)

        for i, (idx, row) in enumerate(df.iterrows()):
            if progress_callback:
                pct = int((i + 1) / total * 100)
                progress_callback(pct, 100, f"Validating address {i + 1}/{total}...")

            parsed = parsed_addresses[i]
            row_fields = self._extract_row_fields(row, parsed, col_map)

            # Skip non-IT countries
            if row_fields["country"] not in ('IT',):
                results.append(ValidationResult(
                    row_index=idx, name=row_fields["name"],
                    city=row_fields["city"], street=row_fields["street"],
                    original_zip=row_fields["original_zip"],
                    suggested_zip=row_fields["original_zip"],
                    confidence=100,
                    reason=f"Non-IT country ({row_fields['country']}) - skipped",
                    is_valid=True, street_verified=True, street_confidence=100,
                    country_code=row_fields["country"],
                    country_detected=row_fields["country_detected"],
                    phone_missing=row_fields["phone_missing"],
                    original_phone=row_fields["original_phone"],
                    cod_changed=row_fields["cod_changed"],
                    original_cod=row_fields["original_cod"],
                ))
                skipped_count += 1
                continue

            result = self._validate_single_row(
                idx, row, parsed, row_fields, col_map, df
            )
            results.append(result)

            # Count stats
            if result.is_valid:
                valid_count += 1
            elif result.auto_corrected:
                corrected_count += 1
            else:
                review_count += 1
            if result.street_verified:
                street_verified_count += 1
            elif result.street_auto_corrected:
                street_corrected_count += 1
            if result.po_invalid:
                po_invalid_count += 1

        logger.info(
            f"Validation complete: {valid_count} valid, {corrected_count} corrected, "
            f"{review_count} need review, {skipped_count} skipped"
        )

        report = ValidationReport(
            total_rows=total,
            valid_count=valid_count,
            corrected_count=corrected_count,
            review_count=review_count,
            skipped_count=skipped_count,
            results=results,
            street_verified_count=street_verified_count,
            street_corrected_count=street_corrected_count,
            po_invalid_count=po_invalid_count
        )
        return report, df

    def _extract_row_fields(self, row, parsed: ParsedAddress, col_map: dict) -> dict:
        """Extract and normalize all fields from a DataFrame row."""
        name = str(row.get(col_map.get('name', ''), ''))
        street = str(row.get(col_map.get('street', ''), ''))
        city = str(row.get(col_map['city'], ''))
        original_zip = str(row.get(col_map['zip'], ''))

        # State/province
        state = ''
        if col_map.get('state'):
            state_raw = row.get(col_map['state'], '')
            state = str(state_raw).strip() if pd.notna(state_raw) else ''

        # Phone
        original_phone = ''
        phone_missing = False
        phone_col = col_map.get('phone')
        if phone_col:
            phone_raw = row.get(phone_col, '')
            original_phone = str(phone_raw).strip() if pd.notna(phone_raw) else ''
            phone_missing = not original_phone or original_phone.lower() == 'nan'

        # Cash on Delivery
        original_cod = ''
        cod_changed = False
        cod_col = col_map.get('cash_on_delivery')
        if cod_col:
            cod_raw = row.get(cod_col, '')
            original_cod = str(cod_raw).strip() if pd.notna(cod_raw) else ''
            try:
                cod_changed = float(original_cod) != 0.0
            except (ValueError, TypeError):
                cod_changed = bool(original_cod) and original_cod.lower() != 'nan'

        # PO validation
        po_value = ''
        po_extracted = ''
        po_invalid = False
        order_col = col_map.get('order_number')
        if order_col:
            order_raw = row.get(order_col, '')
            po_value = str(order_raw).strip() if pd.notna(order_raw) else ''
            if po_value and po_value.lower() != 'nan':
                po_valid, po_extracted, _ = self.validate_po_number(po_value)
                po_invalid = not po_valid

        # Country detection
        country = parsed.country_code
        country_detected = True
        if col_map.get('country'):
            country_raw = row.get(col_map['country'], '')
            country_explicit = str(country_raw).strip() if pd.notna(country_raw) and str(country_raw).strip() else ''
            if country_explicit:
                country_upper = country_explicit.upper()
                country_map = {
                    'ITALY': 'IT', 'ITALIA': 'IT',
                    'GERMANY': 'DE', 'DEUTSCHLAND': 'DE',
                    'FRANCE': 'FR',
                    'SPAIN': 'ES', 'ESPAÑA': 'ES', 'ESPANA': 'ES',
                    'UNITED KINGDOM': 'GB', 'UK': 'GB', 'GREAT BRITAIN': 'GB',
                }
                if country_upper in country_map:
                    country = country_map[country_upper]
                elif len(country_explicit) == 2:
                    country = country_upper
                country_detected = False

        # Street 2
        street2 = ''
        if col_map.get('street2'):
            s2_raw = row.get(col_map['street2'], '')
            street2 = str(s2_raw).strip().rstrip('-').strip() if pd.notna(s2_raw) else ''
            if street2.lower() == 'nan':
                street2 = ''

        return {
            "name": name, "street": street, "city": city,
            "original_zip": original_zip, "state": state,
            "original_phone": original_phone, "phone_missing": phone_missing,
            "original_cod": original_cod, "cod_changed": cod_changed,
            "po_value": po_value, "po_extracted": po_extracted, "po_invalid": po_invalid,
            "country": country, "country_detected": country_detected,
            "street2": street2,
        }

    def _validate_single_row(
        self, idx, row, parsed: ParsedAddress,
        fields: dict, col_map: dict, df: pd.DataFrame,
    ) -> ValidationResult:
        """Validate a single Italian address row against Google API and local DB."""
        city = fields["city"]
        original_zip = fields["original_zip"]
        state = fields["state"]
        street = fields["street"]

        # Pad ZIP
        try:
            zip_padded = str(int(float(str(original_zip)))).zfill(5)
        except (ValueError, TypeError):
            zip_padded = str(original_zip).strip()

        # Call Google API
        api_response = self.address_validator.validate_address(
            parsed, city, zip_padded, state, street2=fields["street2"]
        )

        if not api_response or "result" not in api_response:
            return ValidationResult(
                row_index=idx, name=fields["name"],
                city=city, street=street,
                original_zip=original_zip, suggested_zip=None,
                confidence=0, reason="Google API unavailable",
                is_valid=False,
                country_code=fields["country"],
                country_detected=fields["country_detected"],
                phone_missing=fields["phone_missing"],
                original_phone=fields["original_phone"],
                cod_changed=fields["cod_changed"],
                original_cod=fields["original_cod"],
                po_invalid=fields["po_invalid"],
                po_value=fields["po_value"],
                po_extracted=fields["po_extracted"],
            )

        # Interpret Google verdict
        api_result = api_response["result"]
        outcome = self.address_validator.interpret_verdict(
            api_result.get("verdict", {}),
            api_result.get("address", {}),
            parsed, zip_padded, city
        )

        # Cross-check ZIP with Italian comuni database
        self._cross_check_zip(outcome, zip_padded, city, state)

        # Build suggested street
        suggested_street = None
        if outcome.street_corrected or outcome.silent_correction:
            suggested_street = f"{outcome.output_street} {parsed.house_number}".strip()

        # Map outcome to result fields
        is_valid = outcome.status == "valid"
        auto_corrected = outcome.status == "corrected"
        street_verified = outcome.street_confirmed and not outcome.street_corrected
        street_auto_corrected = outcome.street_corrected

        if is_valid:
            confidence = 100
        elif auto_corrected:
            confidence = 95
        else:
            confidence = 50

        reason = "; ".join(outcome.reasons) if outcome.reasons else outcome.action

        # Street 2 handling moved to generate_corrected_excel — single owner
        # of Street 2 logic (dedup, truncation). _update_street2 is no longer
        # called during validation to avoid double mutation.

        return ValidationResult(
            row_index=idx, name=fields["name"],
            city=city, street=street,
            original_zip=original_zip,
            suggested_zip=outcome.output_zip if outcome.output_zip else None,
            confidence=confidence, reason=reason,
            is_valid=is_valid, auto_corrected=auto_corrected,
            street_verified=street_verified,
            suggested_street=suggested_street,
            street_confidence=95 if street_verified else (85 if street_auto_corrected else 0),
            street_auto_corrected=street_auto_corrected,
            country_code=fields["country"],
            country_detected=fields["country_detected"],
            phone_missing=fields["phone_missing"],
            original_phone=fields["original_phone"],
            cod_changed=fields["cod_changed"],
            original_cod=fields["original_cod"],
            location_info=outcome.location_info or parsed.location_info or "",
            po_invalid=fields["po_invalid"],
            po_value=fields["po_value"],
            po_extracted=fields["po_extracted"],
        )

    def _cross_check_zip(self, outcome: ValidationOutcome, zip_padded: str,
                          city: str, state: str):
        """Cross-check ZIP against Italian comuni database, mutating outcome in place."""
        check_zip = outcome.output_zip or zip_padded
        if not check_zip:
            return

        if not self.address_validator.is_valid_italian_cap(check_zip) and not outcome.zip_confirmed:
            outcome.status = "review"
            outcome.reasons.append(f"CAP {check_zip} not in local database")
            return

        comune_valid, comune_msg = self.address_validator.validate_zip_comune(
            check_zip, city, state
        )
        if not comune_valid and "not in database" not in comune_msg:
            if outcome.status in ("valid", "corrected") and not outcome.zip_confirmed:
                outcome.status = "review"
            outcome.reasons.append(comune_msg)
        elif state:
            prov_valid, prov_msg = self.address_validator.validate_zip_province(
                check_zip, state
            )
            if not prov_valid:
                outcome.status = "review"
                outcome.reasons.append(prov_msg)

    MAX_STREET2_LENGTH = 35  # ShippyPro rejects Street 2 over 35 chars

    # =========================================================================
    # Excel output
    # =========================================================================

    def generate_corrected_excel(
        self,
        original_df: pd.DataFrame,
        report: ValidationReport,
        brand: str = "",
        campaign: str = "",
        po_number: str = "",
    ) -> bytes:
        """Generate corrected Excel with auto-corrections applied."""
        logger.info(f"Generating corrected Excel with {len(report.results)} rows")

        df = original_df.copy()
        col_map = self._map_columns(df)
        zip_col = col_map.get('zip')
        street_col = col_map.get('street')
        city_col = col_map.get('city')
        country_col = col_map.get('country')
        phone_col = col_map.get('phone')
        cod_col = col_map.get('cash_on_delivery')

        for result in report.results:
            # Correct ZIP
            if result.auto_corrected and result.suggested_zip and zip_col:
                df.at[result.row_index, zip_col] = result.suggested_zip

            # Correct street (house number already correct from ParsedAddress)
            if result.street_auto_corrected and result.suggested_street and street_col:
                df.at[result.row_index, street_col] = result.suggested_street

            # Fill country code if empty
            if country_col:
                current_country = df.at[result.row_index, country_col]
                if not str(current_country).strip() or str(current_country).lower() == 'nan':
                    df.at[result.row_index, country_col] = result.country_code

            # Fill phone with default if empty
            if phone_col:
                current_phone = df.at[result.row_index, phone_col]
                if pd.isna(current_phone) or not str(current_phone).strip() or str(current_phone).lower() == 'nan':
                    df.at[result.row_index, phone_col] = self.DEFAULT_PHONE

            # Handle Street 2: move location info, remove from Street 1, truncate
            street2_col = col_map.get('street2')
            if street2_col:
                s2 = str(df.at[result.row_index, street2_col]).strip().rstrip('-').strip()
                if s2.lower() == 'nan':
                    s2 = ''

                loc = result.location_info
                if loc:
                    if not s2:
                        new_s2 = loc
                    elif loc.lower() not in s2.lower():
                        new_s2 = f"{s2} - {loc}"
                    else:
                        new_s2 = s2

                    # Truncate to ShippyPro limit (word-boundary aware)
                    if len(new_s2) > self.MAX_STREET2_LENGTH:
                        cut = new_s2[:self.MAX_STREET2_LENGTH]
                        last_space = cut.rfind(' ')
                        new_s2 = cut[:last_space] if last_space > 20 else cut.rstrip()

                    df.at[result.row_index, street2_col] = new_s2

                    # Remove location info from Street 1 to avoid duplication
                    if street_col:
                        s1 = str(df.at[result.row_index, street_col]).strip()
                        if loc in s1:
                            s1 = ' '.join(s1.replace(loc, '').split())
                            if s1:
                                df.at[result.row_index, street_col] = s1
                elif s2:
                    if len(s2) > self.MAX_STREET2_LENGTH:
                        cut = s2[:self.MAX_STREET2_LENGTH]
                        last_space = cut.rfind(' ')
                        s2 = cut[:last_space] if last_space > 20 else cut.rstrip()
                    df.at[result.row_index, street2_col] = s2

            # Set Cash on Delivery to 0 always (IT addresses only)
            if cod_col and result.country_code == 'IT':
                df.at[result.row_index, cod_col] = 0

        # Add Brand / Campaign / PO columns if provided
        if brand:
            df["Brand"] = brand
        if campaign:
            df["Campaign"] = campaign
        if po_number:
            df["PO Number"] = po_number

        # Sanitize string cells to prevent Excel formula injection
        for col in df.columns:
            df[col] = df[col].apply(
                lambda x: sanitize_cell(x) if isinstance(x, str) else x
            )

        return format_excel_output(df, col_map)

    def generate_review_report(self, report: ValidationReport) -> bytes:
        """Generate Excel report for items needing manual review."""
        from openpyxl.utils import get_column_letter

        review_items = []
        for r in report.results:
            if not r.is_valid or not r.street_verified:
                review_items.append({
                    'Row': r.row_index + 2,
                    'Name': r.name,
                    'City': r.city,
                    'Country': f"{r.country_code}{'*' if r.country_detected else ''}",
                    'Original Street': r.street,
                    'Suggested Street': r.suggested_street or '-',
                    'Street Conf.': f"{r.street_confidence}%" if r.street_confidence else '-',
                    'Street Action': (
                        'Verified' if r.street_verified else
                        'Auto-corrected' if r.street_auto_corrected else
                        'Review needed'
                    ),
                    'Original ZIP': r.original_zip,
                    'Suggested ZIP': r.suggested_zip or '-',
                    'ZIP Conf.': f"{r.confidence}%",
                    'ZIP Action': (
                        'Valid' if r.is_valid else
                        'Auto-corrected' if r.auto_corrected else
                        'Review needed'
                    ),
                    'Reason': r.reason
                })

        df = pd.DataFrame(review_items)

        # Sanitize string cells to prevent Excel formula injection
        for col in df.columns:
            df[col] = df[col].apply(
                lambda x: sanitize_cell(x) if isinstance(x, str) else x
            )

        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Review')
            worksheet = writer.sheets['Review']

            for col_idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                    len(col)
                ) + 2
                column_width = min(max_length, 60)
                worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = column_width

        return output.getvalue()
