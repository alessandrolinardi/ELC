"""
Address Validator Module
Validates and corrects addresses (ZIP codes and streets) using Photon API (primary)
with Nominatim API fallback. Photon has no rate limits (~10x faster), while Nominatim
is used as a reliable fallback.
"""

import time
import requests
from dataclasses import dataclass
from typing import Optional, Callable
from io import BytesIO
import re
from difflib import SequenceMatcher

import pandas as pd

from .logging_config import get_logger

# Logger for this module
logger = get_logger(__name__)


@dataclass
class ValidationResult:
    """Result of a single address validation."""
    row_index: int
    name: str
    city: str
    street: str
    original_zip: str
    suggested_zip: Optional[str]
    confidence: int  # 0-100
    reason: str
    is_valid: bool
    auto_corrected: bool = False
    # Street validation fields
    street_verified: bool = False
    suggested_street: Optional[str] = None
    street_confidence: int = 0
    street_auto_corrected: bool = False
    # Country code (ISO 3166-1 alpha-2)
    country_code: str = "IT"
    country_detected: bool = False  # True if country was auto-detected
    # Additional corrections tracking
    phone_missing: bool = False      # True if phone was empty and will be filled
    original_phone: str = ""         # Original phone value
    cod_changed: bool = False        # True if Cash on Delivery will be set to 0
    original_cod: str = ""           # Original COD value


@dataclass
class ValidationReport:
    """Complete validation report."""
    total_rows: int
    valid_count: int
    corrected_count: int
    review_count: int
    skipped_count: int
    results: list[ValidationResult]
    # Street stats
    street_verified_count: int = 0
    street_corrected_count: int = 0


class ZipValidator:
    """
    Validates addresses using Photon API (primary) with Nominatim fallback.
    Validates both ZIP codes and street names.
    """

    # Photon API - primary (no rate limits, ~10x faster)
    PHOTON_URL = "https://photon.komoot.io/api/"
    # Nominatim API - fallback (1 req/sec limit)
    NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
    USER_AGENT = "ELC-AddressValidator/2.0 (shipping label tool)"
    REQUEST_DELAY = 1.1  # seconds between requests (Nominatim only)
    PHOTON_DELAY = 0.1  # minimal delay for Photon (courtesy)

    # Italian CAP ranges for major cities (for additional validation)
    ITALIAN_CAP_RANGES = {
        'roma': ('00100', '00199'),
        'milano': ('20100', '20199'),
        'napoli': ('80100', '80147'),
        'torino': ('10100', '10156'),
        'firenze': ('50100', '50145'),
        'florence': ('50100', '50145'),
        'bologna': ('40100', '40141'),
        'venezia': ('30100', '30176'),
        'venice': ('30100', '30176'),
    }

    # Italian province codes to ZIP prefix mapping
    # Province code -> list of valid ZIP prefixes (first 2-3 digits)
    ITALIAN_PROVINCE_ZIP = {
        # Piemonte
        'TO': ['10'], 'VC': ['13'], 'NO': ['28'], 'CN': ['12'], 'AT': ['14'],
        'AL': ['15'], 'BI': ['13'], 'VB': ['28'],
        # Valle d'Aosta
        'AO': ['11'],
        # Lombardia
        'VA': ['21'], 'CO': ['22'], 'SO': ['23'], 'MI': ['20'], 'BG': ['24'],
        'BS': ['25'], 'PV': ['27'], 'CR': ['26'], 'MN': ['46'], 'LC': ['23'],
        'LO': ['26'], 'MB': ['20'],
        # Trentino-Alto Adige
        'BZ': ['39'], 'TN': ['38'],
        # Veneto
        'VR': ['37'], 'VI': ['36'], 'BL': ['32'], 'TV': ['31'], 'VE': ['30'],
        'PD': ['35'], 'RO': ['45'],
        # Friuli-Venezia Giulia
        'UD': ['33'], 'GO': ['34'], 'TS': ['34'], 'PN': ['33'],
        # Liguria
        'IM': ['18'], 'SV': ['17'], 'GE': ['16'], 'SP': ['19'],
        # Emilia-Romagna
        'PC': ['29'], 'PR': ['43'], 'RE': ['42'], 'MO': ['41'], 'BO': ['40'],
        'FE': ['44'], 'RA': ['48'], 'FC': ['47'], 'RN': ['47'],
        # Toscana
        'MS': ['54'], 'LU': ['55'], 'PT': ['51'], 'FI': ['50'], 'LI': ['57'],
        'PI': ['56'], 'AR': ['52'], 'SI': ['53'], 'GR': ['58'], 'PO': ['59'],
        # Umbria
        'PG': ['06'], 'TR': ['05'],
        # Marche
        'PU': ['61'], 'AN': ['60'], 'MC': ['62'], 'AP': ['63'], 'FM': ['63'],
        # Lazio
        'VT': ['01'], 'RI': ['02'], 'RM': ['00'], 'LT': ['04'], 'FR': ['03'],
        # Abruzzo
        'AQ': ['67'], 'TE': ['64'], 'PE': ['65'], 'CH': ['66'],
        # Molise
        'CB': ['86'], 'IS': ['86'],
        # Campania
        'CE': ['81'], 'BN': ['82'], 'NA': ['80'], 'AV': ['83'], 'SA': ['84'],
        # Puglia
        'FG': ['71'], 'BA': ['70'], 'TA': ['74'], 'BR': ['72'], 'LE': ['73'],
        'BT': ['76'],
        # Basilicata
        'PZ': ['85'], 'MT': ['75'],
        # Calabria
        'CS': ['87'], 'CZ': ['88'], 'RC': ['89'], 'KR': ['88'], 'VV': ['89'],
        # Sicilia
        'TP': ['91'], 'PA': ['90'], 'ME': ['98'], 'AG': ['92'], 'CL': ['93'],
        'EN': ['94'], 'CT': ['95'], 'RG': ['97'], 'SR': ['96'],
        # Sardegna
        'SS': ['07'], 'NU': ['08'], 'CA': ['09'], 'OR': ['09'], 'SU': ['09'],
    }

    # Common Italian street prefixes for normalization
    STREET_PREFIXES = [
        'via', 'viale', 'piazza', 'piazzale', 'corso', 'largo', 'vicolo',
        'strada', 'contrada', 'borgata', 'traversa', 'salita', 'discesa',
        'lungomare', 'lungotevere', 'lungarno', 'circonvallazione'
    ]

    # Location prefixes that should be moved to Street 2 (Centro Commerciale, etc.)
    LOCATION_PREFIXES = [
        'centro commerciale', 'c.c.', 'cc ', 'c/c',
        'centro direzionale', 'c.d.', 'cd ',
        'centro servizi', 'c.s.',
        'parco commerciale', 'p.c.',
        'galleria commerciale',
        'outlet',
        'retail park',
    ]

    def __init__(self, confidence_threshold: int = 90, street_confidence_threshold: int = 85):
        """
        Initialize validator.

        Args:
            confidence_threshold: Minimum confidence for auto-correction of ZIP (default 90%)
            street_confidence_threshold: Minimum confidence for auto-correction of street (default 85%)
        """
        self.confidence_threshold = confidence_threshold
        self.street_confidence_threshold = street_confidence_threshold
        self.session = requests.Session()
        self.session.headers.update({'User-Agent': self.USER_AGENT})
        self._photon_available = True  # Track if Photon is responding
        self._city_cache = {}  # Cache for city lookups: city_name -> result

    def _query_photon(self, query: str, limit: int = 5) -> list[dict]:
        """
        Query Photon API for address.

        Photon returns GeoJSON format. This method converts it to a format
        compatible with our Nominatim-based processing.

        Args:
            query: Free-text address query (e.g., "Via Roma 10, Milano, Italy")
            limit: Maximum results to return

        Returns:
            List of results in Nominatim-compatible format, or empty list on failure
        """
        if not self._photon_available:
            return []

        try:
            params = {
                'q': query,
                'limit': limit,
                'lang': 'it'  # Italian language for better results
            }

            logger.debug(f"Photon query: '{query}'")
            response = self.session.get(self.PHOTON_URL, params=params, timeout=5)
            response.raise_for_status()
            data = response.json()

            # Convert GeoJSON features to Nominatim-like format
            results = []
            for feature in data.get('features', []):
                props = feature.get('properties', {})

                # Build address dict matching Nominatim structure
                address = {
                    'road': props.get('street'),
                    'house_number': props.get('housenumber'),
                    'postcode': props.get('postcode'),
                    'city': props.get('city'),
                    'town': props.get('city'),  # Photon uses 'city' for both
                    'village': props.get('city'),
                    'state': props.get('state'),
                    'country': props.get('country'),
                    'country_code': props.get('countrycode', '').upper() if props.get('countrycode') else None
                }

                # Create Nominatim-compatible result
                result = {
                    'display_name': props.get('name', ''),
                    'address': address,
                    'lat': feature.get('geometry', {}).get('coordinates', [0, 0])[1],
                    'lon': feature.get('geometry', {}).get('coordinates', [0, 0])[0],
                    '_source': 'photon'  # Track source
                }
                results.append(result)

            if results:
                logger.debug(f"Photon found {len(results)} results")
            return results

        except requests.exceptions.Timeout:
            logger.warning("Photon API timeout - will use Nominatim fallback")
            self._photon_available = False
            return []
        except requests.exceptions.HTTPError as e:
            # 400 errors are likely bad input, don't disable Photon
            if e.response is not None and e.response.status_code == 400:
                logger.debug(f"Photon 400 error for query '{query}' - skipping")
                return []
            # Other HTTP errors might indicate service issues
            logger.warning(f"Photon API error: {e} - will use Nominatim fallback")
            self._photon_available = False
            return []
        except requests.exceptions.RequestException as e:
            logger.warning(f"Photon API error: {e} - will use Nominatim fallback")
            self._photon_available = False
            return []
        except Exception as e:
            logger.debug(f"Photon query failed: {e}")
            return []

    def _query_address(self, street: str, city: str, country: str = "Italy") -> Optional[dict]:
        """
        Query address using hybrid approach: Photon first, Nominatim fallback.

        Args:
            street: Street address
            city: City name
            country: Country name

        Returns:
            First matching result or None
        """
        def find_best_result(results: list, city_name: str) -> Optional[dict]:
            """Find the best matching result from a list."""
            best_result = None
            best_with_street = None

            for result in results:
                address = result.get('address', {})
                result_city = (address.get('city') or address.get('town') or '').lower()

                # Check if result matches our city
                if city_name.lower() in result_city or result_city in city_name.lower():
                    if address.get('road'):
                        best_with_street = result
                        break
                    elif best_result is None:
                        best_result = result

            if best_with_street:
                return best_with_street
            elif best_result:
                best_result['_city_only'] = True
                return best_result
            return None

        # Try Photon first (fast, no rate limits)
        if self._photon_available:
            # Build query string for Photon
            parts = [p for p in [street, city, country] if p and p.strip()]
            query = ", ".join(parts)

            results = self._query_photon(query, limit=5)
            if results:
                best = find_best_result(results, city)
                if best and best.get('address', {}).get('postcode'):
                    return best

            # If no good result, try without house number (e.g., "11/A" or "123")
            # This helps with addresses like "Piazza Marescotti 11/A"
            if street:
                street_no_num = re.sub(r'\s*\d+[/\-]?\w*\s*$', '', street).strip()
                if street_no_num and street_no_num != street:
                    parts = [p for p in [street_no_num, city, country] if p and p.strip()]
                    query = ", ".join(parts)
                    results = self._query_photon(query, limit=5)
                    if results:
                        best = find_best_result(results, city)
                        if best and best.get('address', {}).get('postcode'):
                            return best

            # If still no result, try extracting main street from complex addresses
            # e.g., "C.C. Le Grange Via Casilina inc. Via Marello 1" -> "Via Casilina"
            if street:
                # Look for first occurrence of a street prefix
                street_lower = street.lower()
                for prefix in self.STREET_PREFIXES:
                    match = re.search(rf'\b({prefix}\.?\s+\w+)', street_lower)
                    if match:
                        # Extract the main street (prefix + first word)
                        start = match.start()
                        # Get from original string to preserve case
                        simple_street = street[start:]
                        # Cut at "inc.", intersection markers, or next street prefix
                        simple_street = re.split(r'\s+(?:inc\.?|incr\.?|ang\.?|angolo)\s+', simple_street, maxsplit=1)[0]
                        simple_street = re.sub(r'\s*\d+[/\-]?\w*\s*$', '', simple_street).strip()

                        if simple_street and simple_street.lower() != street.lower():
                            parts = [p for p in [simple_street, city, country] if p and p.strip()]
                            query = ", ".join(parts)
                            results = self._query_photon(query, limit=5)
                            if results:
                                best = find_best_result(results, city)
                                if best and best.get('address', {}).get('postcode'):
                                    return best
                        break

            # Return whatever we have from initial query
            if results:
                best = find_best_result(results, city)
                if best:
                    return best
                # Fall back to first result
                result = results[0]
                if not result.get('address', {}).get('road'):
                    result['_city_only'] = True
                return result

            # No results from Photon - try city-only lookup from cache or quick query
            city_lower = city.lower().strip()
            if city_lower in self._city_cache:
                return self._city_cache[city_lower]

            # Try city-only query with Photon (fast)
            city_query = f"{city}, {country}"
            city_results = self._query_photon(city_query, limit=3)
            if city_results:
                for cr in city_results:
                    if cr.get('address', {}).get('postcode'):
                        cr['_city_only'] = True
                        self._city_cache[city_lower] = cr
                        return cr

            # Cache the miss to avoid repeated queries
            self._city_cache[city_lower] = None
            return None

        # Only use Nominatim when Photon service is completely down
        logger.debug("Using Nominatim fallback (Photon unavailable)")
        return self._query_nominatim(street, city, country)

    def detect_country_code(self, zip_code: str, city: str = "", street: str = "") -> str:
        """
        Detect country code (ISO 3166-1 alpha-2) from address components.

        Args:
            zip_code: ZIP/postal code
            city: City name (optional, helps with ambiguous cases)
            street: Street name (optional, helps detect language)

        Returns:
            2-letter country code (e.g., 'IT', 'DE', 'FR'). Defaults to 'IT' if uncertain.
        """
        zip_clean = re.sub(r'[^A-Z0-9]', '', str(zip_code).upper().strip())
        city_lower = city.lower().strip() if city else ""
        street_lower = street.lower().strip() if street else ""

        # UK: Alphanumeric format like "SW1A1AA", "EC1A1BB", "M11AA"
        if re.match(r'^[A-Z]{1,2}[0-9][0-9A-Z]?\s*[0-9][A-Z]{2}$', zip_clean):
            return 'GB'

        # Netherlands: 4 digits + 2 letters (1234AB)
        if re.match(r'^\d{4}[A-Z]{2}$', zip_clean):
            return 'NL'

        # Portugal: 4 digits - 3 digits (1234-567 or 1234567)
        if re.match(r'^\d{7}$', zip_clean) or re.match(r'^\d{4}-?\d{3}$', str(zip_code)):
            return 'PT'

        # 5-digit ZIP codes - need to distinguish by range and context
        if re.match(r'^\d{5}$', zip_clean):
            first_two = int(zip_clean[:2])
            first_digit = int(zip_clean[0])

            # Italy: 00xxx-99xxx, but check for Italian cities/streets
            italian_cities = ['roma', 'milano', 'napoli', 'torino', 'firenze', 'bologna',
                              'venezia', 'palermo', 'genova', 'bari', 'catania', 'verona',
                              'padova', 'trieste', 'brescia', 'parma', 'modena', 'reggio',
                              'ravenna', 'ferrara', 'rimini', 'livorno', 'cagliari', 'sassari',
                              'perugia', 'ancona', 'pescara', 'trento', 'bolzano', 'aosta']

            italian_prefixes = ['via ', 'viale ', 'piazza ', 'corso ', 'largo ', 'vicolo ',
                                'strada ', 'piazzale ', 'lungomare ', 'circonvallazione ']

            # Check for Italian indicators
            is_italian_city = any(it_city in city_lower for it_city in italian_cities)
            is_italian_street = any(street_lower.startswith(prefix) for prefix in italian_prefixes)

            if is_italian_city or is_italian_street:
                return 'IT'

            # Germany: 01xxx-99xxx (overlaps with Italy, use city/street clues)
            german_cities = ['berlin', 'hamburg', 'münchen', 'munich', 'köln', 'cologne',
                             'frankfurt', 'stuttgart', 'düsseldorf', 'dortmund', 'essen',
                             'leipzig', 'bremen', 'dresden', 'hannover', 'nürnberg']
            german_prefixes = ['straße', 'strasse', 'str.', 'platz', 'weg', 'allee', 'ring']

            is_german_city = any(de_city in city_lower for de_city in german_cities)
            is_german_street = any(prefix in street_lower for prefix in german_prefixes)

            if is_german_city or is_german_street:
                return 'DE'

            # France: 01xxx-98xxx
            french_cities = ['paris', 'marseille', 'lyon', 'toulouse', 'nice', 'nantes',
                             'strasbourg', 'montpellier', 'bordeaux', 'lille', 'rennes']
            french_prefixes = ['rue ', 'avenue ', 'boulevard ', 'place ', 'allée ', 'chemin ']

            is_french_city = any(fr_city in city_lower for fr_city in french_cities)
            is_french_street = any(street_lower.startswith(prefix) for prefix in french_prefixes)

            if is_french_city or is_french_street:
                return 'FR'

            # Spain: 01xxx-52xxx
            if first_two <= 52:
                spanish_cities = ['madrid', 'barcelona', 'valencia', 'sevilla', 'zaragoza',
                                  'málaga', 'malaga', 'murcia', 'palma', 'bilbao']
                spanish_prefixes = ['calle ', 'avenida ', 'plaza ', 'paseo ', 'carrer ']

                is_spanish_city = any(es_city in city_lower for es_city in spanish_cities)
                is_spanish_street = any(street_lower.startswith(prefix) for prefix in spanish_prefixes)

                if is_spanish_city or is_spanish_street:
                    return 'ES'

            # Default to Italy for 5-digit codes (most common use case for this app)
            return 'IT'

        # Austria: 4 digits (1xxx-9xxx)
        if re.match(r'^\d{4}$', zip_clean):
            austrian_cities = ['wien', 'vienna', 'graz', 'linz', 'salzburg', 'innsbruck']
            if any(at_city in city_lower for at_city in austrian_cities):
                return 'AT'

            # Could also be Switzerland, Belgium, etc. - check context
            swiss_cities = ['zürich', 'zurich', 'genève', 'geneva', 'basel', 'bern', 'lausanne']
            if any(ch_city in city_lower for ch_city in swiss_cities):
                return 'CH'

            # Default to Austria for 4-digit
            return 'AT'

        # If no pattern matches, default to Italy
        return 'IT'

    def _normalize_street(self, street: str) -> str:
        """
        Normalize street name for comparison.
        Removes common prefixes and standardizes format.
        """
        if not street:
            return ""

        normalized = street.lower().strip()

        # Remove common prefixes for comparison
        for prefix in self.STREET_PREFIXES:
            if normalized.startswith(prefix + ' '):
                normalized = normalized[len(prefix) + 1:]
                break
            if normalized.startswith(prefix + '.'):
                normalized = normalized[len(prefix) + 1:]
                break

        # Remove punctuation and extra spaces
        normalized = re.sub(r'[.,;:]', ' ', normalized)
        normalized = re.sub(r'\s+', ' ', normalized).strip()

        # Remove house numbers at the end for comparison (e.g., "21 21", "11/A", "123")
        normalized = re.sub(r'\s+\d+[/\-]?\w*(\s+\d+[/\-]?\w*)*\s*$', '', normalized).strip()

        return normalized

    def _extract_street_name(self, street: str) -> tuple[str, str]:
        """
        Extract street prefix and name separately.

        Returns:
            Tuple of (prefix, name)
        """
        if not street:
            return "", ""

        street_lower = street.lower().strip()

        for prefix in self.STREET_PREFIXES:
            if street_lower.startswith(prefix + ' '):
                return prefix, street[len(prefix) + 1:].strip()
            if street_lower.startswith(prefix + '.'):
                return prefix, street[len(prefix) + 1:].strip()

        return "", street

    def _extract_house_number(self, street: str) -> tuple[str, str]:
        """
        Extract house number from the end of a street address.

        Returns:
            Tuple of (street_without_number, house_number)
        """
        if not street:
            return "", ""

        # Match house numbers at the end: "21", "21 21", "11/A", "123bis", etc.
        match = re.search(r'\s+(\d+[/\-]?\w*(?:\s+\d+[/\-]?\w*)*)\s*$', street)
        if match:
            house_num = match.group(1)
            street_only = street[:match.start()].strip()
            return street_only, house_num

        return street, ""

    def _build_street_suggestion(self, suggested_name: str, original_street: str) -> str:
        """
        Build a street suggestion preserving the original house number.

        Args:
            suggested_name: The corrected street name from API
            original_street: The original street with house number

        Returns:
            Suggested street with original house number appended
        """
        _, house_num = self._extract_house_number(original_street)
        if house_num:
            return f"{suggested_name} {house_num}"
        return suggested_name

    def _extract_location_prefix(self, street: str) -> tuple[str, str]:
        """
        Extract location prefix (Centro Commerciale, C.C., etc.) from street address.

        Moves location prefixes to a separate field so the actual street address
        can be validated independently.

        Example:
            Input: "C.C. Le Grange Via Casilina inc. Via Marello 1"
            Output: ("Via Casilina inc. Via Marello 1", "C.C. Le Grange")

        Args:
            street: Original street address

        Returns:
            Tuple of (clean_street, extracted_location)
            If no location prefix found, returns (original_street, "")
        """
        if not street:
            return street, ""

        street_lower = street.lower().strip()

        # Check if the street starts with a location prefix
        for prefix in self.LOCATION_PREFIXES:
            if street_lower.startswith(prefix):
                # Find where the actual street address begins (first STREET_PREFIXES match)
                for street_prefix in self.STREET_PREFIXES:
                    match = re.search(rf'\b{street_prefix}\.?\s+', street_lower)
                    if match:
                        # Extract: everything before street prefix is location, rest is street
                        location = street[:match.start()].strip()
                        clean_street = street[match.start():].strip()
                        return clean_street, location

                # No standard street prefix found - return original
                return street, ""

        return street, ""

    def preprocess_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Preprocess DataFrame by moving location prefixes from Street 1 to Street 2.

        This separates Centro Commerciale, C.C., etc. from the actual street address
        to improve validation accuracy.

        Args:
            df: Original DataFrame

        Returns:
            Preprocessed DataFrame with locations moved to Street 2
        """
        df = df.copy()
        col_map = self._map_columns(df)
        street_col = col_map.get('street')
        street2_col = col_map.get('street2')

        if not street_col:
            return df

        # If Street 2 column doesn't exist, create it
        if not street2_col:
            # Find position after Street 1 to insert Street 2
            if street_col in df.columns:
                col_idx = df.columns.get_loc(street_col) + 1
                df.insert(col_idx, 'Street 2', '')
                street2_col = 'Street 2'
            else:
                df['Street 2'] = ''
                street2_col = 'Street 2'

        # Process each row
        for idx, row in df.iterrows():
            street_val = str(row.get(street_col, '')).strip()
            street2_val = str(row.get(street2_col, '')).strip()

            # Skip if Street 1 is empty or nan
            if not street_val or street_val.lower() == 'nan':
                continue

            # Extract location prefix
            clean_street, location = self._extract_location_prefix(street_val)

            # If we extracted a location and Street 2 is empty, move it
            if location and (not street2_val or street2_val.lower() == 'nan'):
                df.at[idx, street_col] = clean_street
                df.at[idx, street2_col] = location
                logger.debug(f"Row {idx}: Moved '{location}' to Street 2, keeping '{clean_street}' in Street 1")

        return df

    def _looks_like_valid_italian_street(self, street: str) -> bool:
        """
        Check if street follows a valid Italian street format.
        This is a heuristic fallback when API can't verify the street.

        Returns:
            True if street format appears valid
        """
        if not street:
            return False

        street_lower = street.lower().strip()

        # Check if starts with a known Italian street prefix
        for prefix in self.STREET_PREFIXES:
            if street_lower.startswith(prefix + ' ') or street_lower.startswith(prefix + '.'):
                # Must have something after the prefix (the actual street name)
                remaining = street_lower[len(prefix):].strip(' .')
                if len(remaining) >= 2:  # At least 2 chars for a name
                    return True

        return False

    def _string_similarity(self, s1: str, s2: str) -> float:
        """
        Calculate similarity ratio between two strings.

        Returns:
            Similarity score 0.0 to 1.0
        """
        if not s1 or not s2:
            return 0.0

        # Normalize both strings
        n1 = self._normalize_street(s1)
        n2 = self._normalize_street(s2)

        return SequenceMatcher(None, n1, n2).ratio()

    def _clean_zip_code(self, zip_code: str) -> tuple[str, bool]:
        """
        Clean zip code by replacing common typos and padding with leading zeros.

        For Italian CAPs (5 digits), if the ZIP is shorter than 5 digits,
        assumes missing digits are leading zeros (e.g., "187" → "00187").

        Args:
            zip_code: Original zip code

        Returns:
            Tuple of (cleaned_zip, was_cleaned)
        """
        original = str(zip_code).strip()

        # Common letter-to-number replacements
        replacements = {
            'o': '0', 'O': '0',  # o/O → 0
            'l': '1', 'I': '1', 'i': '1',  # l/I/i → 1
            'z': '2', 'Z': '2',  # z/Z → 2
            's': '5', 'S': '5',  # s/S → 5
            'b': '8', 'B': '8',  # b/B → 8
            'g': '9', 'G': '9',  # g/G → 9
        }

        cleaned = original
        for char, replacement in replacements.items():
            cleaned = cleaned.replace(char, replacement)

        # Remove any remaining non-digit characters
        cleaned = re.sub(r'[^\d]', '', cleaned)

        # Pad with leading zeros if shorter than 5 digits (Italian CAP)
        # e.g., "187" → "00187" (Rome), "6100" → "06100" (Perugia)
        if len(cleaned) < 5 and len(cleaned) > 0:
            cleaned = cleaned.zfill(5)

        was_cleaned = cleaned != original.replace(' ', '')
        return cleaned, was_cleaned

    def _count_different_digits(self, zip1: str, zip2: str) -> int:
        """
        Count how many digits are different between two zip codes.
        """
        if len(zip1) != len(zip2):
            return max(len(zip1), len(zip2))

        return sum(1 for a, b in zip(zip1, zip2) if a != b)

    def _is_transposition(self, zip1: str, zip2: str) -> bool:
        """
        Check if two zip codes are transpositions (same digits, different order).
        """
        if len(zip1) != len(zip2):
            return False

        return sorted(zip1) == sorted(zip2)

    def _is_adjacent_swap(self, zip1: str, zip2: str) -> bool:
        """
        Check if two zip codes differ by a single adjacent digit swap.
        """
        if len(zip1) != len(zip2):
            return False

        diff_positions = [i for i in range(len(zip1)) if zip1[i] != zip2[i]]

        if len(diff_positions) != 2:
            return False

        i, j = diff_positions
        if j - i != 1:
            return False

        return zip1[i] == zip2[j] and zip1[j] == zip2[i]

    def _is_valid_italian_zip_format(self, zip_code: str) -> bool:
        """Check if zip code has valid Italian CAP format (5 digits)."""
        return bool(re.match(r'^\d{5}$', str(zip_code).strip()))

    def _search_streets_in_city(self, city: str, limit: int = 50) -> list[dict]:
        """
        Search for streets in a city using Nominatim.

        Args:
            city: City name
            limit: Maximum results to return

        Returns:
            List of address results with street information
        """
        try:
            params = {
                'city': city,
                'country': 'Italy',
                'format': 'json',
                'addressdetails': 1,
                'limit': limit
            }

            response = self.session.get(self.NOMINATIM_URL, params=params, timeout=10)
            response.raise_for_status()
            return response.json()

        except Exception:
            return []

    def _query_nominatim(self, street: str, city: str, country: str = "Italy") -> Optional[dict]:
        """
        Query Nominatim API for address (used as fallback when Photon fails).

        Args:
            street: Street address
            city: City name
            country: Country name

        Returns:
            First result from API or None
        """
        try:
            # First try with full address
            params = {
                'street': street,
                'city': city,
                'country': country,
                'format': 'json',
                'addressdetails': 1,
                'limit': 1
            }

            logger.debug(f"Nominatim query: street='{street}', city='{city}'")
            response = self.session.get(self.NOMINATIM_URL, params=params, timeout=10)
            response.raise_for_status()
            results = response.json()

            if results:
                logger.debug(f"Nominatim found address match")
                result = results[0]
                result['_source'] = 'nominatim'
                return result

            # Fallback: city-only search
            logger.debug(f"No street match, trying city-only for '{city}'")
            time.sleep(self.REQUEST_DELAY)
            params = {
                'city': city,
                'country': country,
                'format': 'json',
                'addressdetails': 1,
                'limit': 1
            }

            response = self.session.get(self.NOMINATIM_URL, params=params, timeout=10)
            response.raise_for_status()
            results = response.json()

            if results:
                result = results[0]
                result['_city_only'] = True  # Mark as city-only match
                result['_source'] = 'nominatim'
                logger.debug(f"Nominatim city-only match found")
                return result

            logger.warning(f"Nominatim: no results for '{city}'")
            return None

        except Exception:
            return None

    def _search_similar_streets(self, street: str, city: str) -> list[tuple[str, float, dict]]:
        """
        Search for streets similar to the input in the given city.
        Uses Photon API only for maximum speed.

        Args:
            street: Street to search for
            city: City name

        Returns:
            List of (street_name, similarity_score, full_result) sorted by similarity
        """
        if not street or not city:
            return []

        matches = []
        seen_streets = set()

        try:
            # Extract street components
            street_prefix, street_name = self._extract_street_name(street)

            # Use Photon API (fast, no rate limits)
            if self._photon_available:
                query = f"{street}, {city}, Italy"
                results = self._query_photon(query, limit=10)

                for result in results:
                    address = result.get('address', {})
                    found_street = address.get('road')

                    if found_street and found_street.lower() not in seen_streets:
                        seen_streets.add(found_street.lower())
                        similarity = self._string_similarity(street, found_street)
                        matches.append((found_street, similarity, result))

                # If no good matches, try alternate query without street number
                if (not matches or max(m[1] for m in matches) < 0.70) and street_name:
                    street_name_no_num = re.sub(r',?\s*\d+.*$', '', street_name).strip()
                    if street_name_no_num:
                        search_term = f"{street_prefix} {street_name_no_num}".strip() if street_prefix else street_name_no_num
                        query = f"{search_term}, {city}, Italy"
                        results = self._query_photon(query, limit=10)

                        for result in results:
                            address = result.get('address', {})
                            found_street = address.get('road')

                            if found_street and found_street.lower() not in seen_streets:
                                seen_streets.add(found_street.lower())
                                similarity = self._string_similarity(street, found_street)
                                matches.append((found_street, similarity, result))

            # Sort by similarity and return whatever we have
            matches.sort(key=lambda x: x[1], reverse=True)
            return matches

        except Exception:
            return matches if matches else []

    def validate_address(
        self,
        street: str,
        city: str,
        original_zip: str,
        country: str = "IT"
    ) -> tuple[bool, Optional[str], int, str, bool, Optional[str], int]:
        """
        Validate a full address (ZIP and street).

        Args:
            street: Street address
            city: City name
            original_zip: Original zip code to validate
            country: Country code

        Returns:
            Tuple of (
                zip_valid, suggested_zip, zip_confidence, zip_reason,
                street_verified, suggested_street, street_confidence
            )
        """
        original_zip_raw = str(original_zip).strip()
        original_street = str(street).strip() if street else ""

        # Default street validation values
        street_verified = False
        suggested_street = None
        street_confidence = 0

        # Only validate Italian addresses for now
        if country.upper() not in ('IT', 'ITALY'):
            return True, original_zip_raw, 100, "Non-IT country - skipped", True, original_street, 100

        # Try to clean up the zip code
        cleaned_zip, was_cleaned = self._clean_zip_code(original_zip_raw)

        # Track if original ZIP was very incomplete (1-3 digits)
        original_digits = re.sub(r'[^\d]', '', original_zip_raw)
        was_incomplete = len(original_digits) < 4

        # Format check on cleaned version
        if not self._is_valid_italian_zip_format(cleaned_zip):
            result = self._query_nominatim(original_street or "", city, "Italy")
            if result:
                suggested = result.get('address', {}).get('postcode')
                if suggested:
                    if ';' in suggested:
                        suggested = suggested.split(';')[0]
                    return False, suggested, 75, f"Invalid format '{original_zip_raw}' - suggested from address", False, None, 0
            return False, None, 0, f"Invalid format '{original_zip_raw}' (must be 5 digits)", False, None, 0

        working_zip = cleaned_zip

        # Query address using hybrid approach (Photon first, then Nominatim)
        result = self._query_address(original_street or "", city, "Italy")

        if not result:
            return False, None, 0, "Address not found in API", False, None, 0

        is_city_only = result.get('_city_only', False)

        # Check street match - always try even if city-only match
        if original_street:
            address_data = result.get('address', {})
            found_street = (
                address_data.get('road') or
                address_data.get('pedestrian') or
                address_data.get('footway') or
                address_data.get('residential')
            )

            # If we have a street from the response (not city-only), compare it
            if found_street and not is_city_only:
                similarity = self._string_similarity(original_street, found_street)

                if similarity >= 0.95:
                    # Excellent match
                    street_verified = True
                    street_confidence = 100
                elif similarity >= 0.85:
                    # Good match - likely minor typo in street name
                    street_verified = True
                    suggested_street = self._build_street_suggestion(found_street, original_street)
                    street_confidence = int(similarity * 100)
                elif similarity >= 0.70:
                    # Moderate match - suggest correction
                    street_verified = False
                    suggested_street = self._build_street_suggestion(found_street, original_street)
                    street_confidence = int(similarity * 100)
                else:
                    # Low match - search for similar streets
                    # Note: _search_similar_streets handles its own rate limiting
                    similar = self._search_similar_streets(original_street, city)
                    if similar:
                        best_match, best_score, _ = similar[0]
                        if best_score >= 0.70:
                            suggested_street = self._build_street_suggestion(best_match, original_street)
                            street_confidence = int(best_score * 100)
                        elif best_score >= 0.60:
                            # Lower threshold for suggestion (not auto-correct)
                            suggested_street = self._build_street_suggestion(best_match, original_street)
                            street_confidence = int(best_score * 100)
            else:
                # No street in response OR city-only match - search for similar streets
                # Note: _search_similar_streets handles its own rate limiting
                similar = self._search_similar_streets(original_street, city)
                if similar:
                    best_match, best_score, best_result = similar[0]
                    if best_score >= 0.85:
                        # High confidence - verify the street
                        street_verified = True
                        suggested_street = self._build_street_suggestion(best_match, original_street) if best_score < 0.95 else None
                        street_confidence = int(best_score * 100)
                    elif best_score >= 0.70:
                        suggested_street = self._build_street_suggestion(best_match, original_street)
                        street_confidence = int(best_score * 100)
                    elif best_score >= 0.60:
                        # Lower threshold for suggestion only
                        suggested_street = self._build_street_suggestion(best_match, original_street)
                        street_confidence = int(best_score * 100)

                    # Also get ZIP from best match if available and we're in city-only mode
                    if best_result and is_city_only and best_score >= 0.70:
                        better_zip = best_result.get('address', {}).get('postcode')
                        if better_zip and ';' not in better_zip:
                            result = best_result
                else:
                    # API didn't find a matching street - check if format looks valid
                    # This prevents showing "-" for correctly formatted streets that OSM doesn't have
                    if self._looks_like_valid_italian_street(original_street):
                        if not is_city_only:
                            # API found the address location, just not the street name
                            street_verified = True
                            street_confidence = 75
                        else:
                            # City-only match with valid street format - low confidence verify
                            street_verified = True
                            street_confidence = 60

        # Now validate ZIP
        suggested_zip = result.get('address', {}).get('postcode')

        if not suggested_zip:
            # Try city-only query to get postal code for the city
            city_query = f"{city}, Italy"
            city_results = self._query_photon(city_query, limit=3)
            for city_result in city_results:
                city_zip = city_result.get('address', {}).get('postcode')
                if city_zip and ';' not in city_zip:
                    suggested_zip = city_zip
                    is_city_only = True
                    break

        if not suggested_zip:
            city_lower = city.lower().strip()
            if city_lower in self.ITALIAN_CAP_RANGES:
                cap_start, cap_end = self.ITALIAN_CAP_RANGES[city_lower]
                if cap_start <= working_zip <= cap_end:
                    return False, working_zip, 70, f"Street not found - ZIP in city range ({cap_start}-{cap_end})", street_verified, suggested_street, street_confidence
                else:
                    return False, cap_start, 80, f"ZIP outside city range ({cap_start}-{cap_end})", street_verified, suggested_street, street_confidence
            # No postal code from API - return the cleaned/padded ZIP if it was modified
            if was_cleaned:
                return False, working_zip, 60, f"No API postal code - using padded ZIP '{original_zip_raw}' → '{working_zip}'", street_verified, suggested_street, street_confidence
            return False, working_zip, 50, "No postal code in API response", street_verified, suggested_street, street_confidence

        # Handle multiple postcodes
        if ';' in suggested_zip:
            suggested_zips = suggested_zip.split(';')
            if working_zip in suggested_zips:
                return True, working_zip, 100, "Exact match (one of multiple)", street_verified, suggested_street, street_confidence
            for sz in suggested_zips:
                if self._count_different_digits(working_zip, sz) == 1:
                    suggested_zip = sz
                    break
            else:
                suggested_zip = suggested_zips[0]

        # Exact match
        if working_zip == suggested_zip:
            if was_cleaned:
                return False, suggested_zip, 95, f"Typo fixed: '{original_zip_raw}' → '{suggested_zip}'", street_verified, suggested_street, street_confidence
            return True, working_zip, 100, "Exact match", street_verified, suggested_street, street_confidence

        # Calculate confidence
        diff_count = self._count_different_digits(working_zip, suggested_zip)
        is_transposition = self._is_transposition(working_zip, suggested_zip)
        is_adjacent_swap = self._is_adjacent_swap(working_zip, suggested_zip)

        # If original ZIP was incomplete (1-3 digits), suggest the padded version
        # not the API's suggestion - flag for manual review
        if was_incomplete and diff_count >= 2:
            confidence = 50
            reason = f"Original ZIP '{original_zip_raw}' padded to '{working_zip}' - needs manual review"
            return False, working_zip, confidence, reason, street_verified, suggested_street, street_confidence
        elif is_city_only:
            confidence = 70
            reason = "City-level match only (street not found)"
        elif is_adjacent_swap:
            confidence = 96
            reason = f"Adjacent digits swapped ({working_zip} → {suggested_zip})"
        elif is_transposition:
            confidence = 95
            reason = f"Digits transposed ({working_zip} → {suggested_zip})"
        elif diff_count == 1:
            confidence = 95
            reason = f"Typo: 1 digit different ({working_zip} → {suggested_zip})"
        elif diff_count == 2:
            confidence = 92
            reason = f"2 digits different ({working_zip} → {suggested_zip})"
        elif diff_count >= 3 and not is_city_only:
            # 3+ digits different is suspicious
            # Check if original ZIP looks correct for the city (same province, city center ZIP)
            same_province = working_zip[:2] == suggested_zip[:2]
            is_city_center_zip = working_zip.endswith('100') or working_zip.endswith('00')

            if same_province and is_city_center_zip:
                # Original ZIP looks like a valid city center ZIP in the same province
                # API probably found the street in a different town - trust original
                confidence = 85
                reason = f"Original ZIP {working_zip} appears correct (city center) - API found {suggested_zip}"
                return False, working_zip, confidence, reason, street_verified, suggested_street, street_confidence
            else:
                confidence = 70
                reason = f"{diff_count} digits different - verify manually ({working_zip} → {suggested_zip})"
        else:
            confidence = 85
            reason = f"{diff_count} digits different - verify manually"

        if was_cleaned and not was_incomplete:
            reason = f"Cleaned '{original_zip_raw}' → '{working_zip}'. " + reason

        return False, suggested_zip, confidence, reason, street_verified, suggested_street, street_confidence

    def validate_zip(
        self,
        street: str,
        city: str,
        original_zip: str,
        country: str = "IT"
    ) -> tuple[bool, Optional[str], int, str]:
        """
        Validate a single zip code (backwards compatible method).

        Returns:
            Tuple of (is_valid, suggested_zip, confidence, reason)
        """
        zip_valid, suggested_zip, confidence, reason, _, _, _ = self.validate_address(
            street, city, original_zip, country
        )
        return zip_valid, suggested_zip, confidence, reason

    def process_dataframe(
        self,
        df: pd.DataFrame,
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
        preprocess: bool = True
    ) -> tuple[ValidationReport, pd.DataFrame]:
        """
        Process entire DataFrame and validate all addresses.

        Args:
            df: DataFrame with address data
            progress_callback: Optional callback(current, total, message)
            preprocess: If True, preprocess to move Centro Commerciale to Street 2

        Returns:
            Tuple of (ValidationReport with all results, preprocessed DataFrame)
        """
        logger.info(f"Starting address validation for {len(df)} rows")

        # Preprocess to move Centro Commerciale info to Street 2
        if preprocess:
            logger.info("Preprocessing: moving location prefixes to Street 2")
            df = self.preprocess_dataframe(df)

        results = []
        valid_count = 0
        corrected_count = 0
        review_count = 0
        skipped_count = 0
        street_verified_count = 0
        street_corrected_count = 0

        col_map = self._map_columns(df)
        logger.debug(f"Column mapping: {col_map}")

        # Only city and zip are required; country can be auto-detected
        if not all(col_map.get(k) for k in ['city', 'zip']):
            raise ValueError(
                f"Missing required columns. Found: {list(df.columns)}\n"
                f"Need: City, Zip (Country is optional - will be auto-detected)"
            )

        has_country_col = col_map.get('country') is not None
        has_state_col = col_map.get('state') is not None
        total = len(df)

        for idx, row in df.iterrows():
            if progress_callback:
                progress_callback(idx + 1, total, f"Validating row {idx + 1}...")

            name = str(row.get(col_map.get('name', ''), ''))
            street = str(row.get(col_map.get('street', ''), ''))
            # Combine Street 1 and Street 2 if both present
            street2 = str(row.get(col_map.get('street2', ''), '')) if col_map.get('street2') else ''
            if street2 and street2.lower() != 'nan':
                full_street = f"{street}, {street2}".strip(', ')
            else:
                full_street = street
            city = str(row.get(col_map['city'], ''))
            original_zip = str(row.get(col_map['zip'], ''))
            # Get state/province if available (for ZIP validation)
            state = ''
            if has_state_col:
                state_raw = row.get(col_map['state'], '')
                state = str(state_raw).strip() if pd.notna(state_raw) else ''

            # Get phone if available - track if missing
            phone_col = col_map.get('phone')
            original_phone = ''
            phone_missing = False
            if phone_col:
                phone_raw = row.get(phone_col, '')
                original_phone = str(phone_raw).strip() if pd.notna(phone_raw) else ''
                phone_missing = not original_phone or original_phone.lower() == 'nan'

            # Get Cash on Delivery if available - track if needs to be set to 0
            cod_col = col_map.get('cash_on_delivery')
            original_cod = ''
            cod_changed = False
            if cod_col:
                cod_raw = row.get(cod_col, '')
                original_cod = str(cod_raw).strip() if pd.notna(cod_raw) else ''
                # COD needs to be changed if it's not 0
                cod_changed = original_cod != '0' and original_cod.lower() != 'nan' and original_cod != ''

            # Get country from column or auto-detect
            country_detected = False
            if has_country_col:
                country_raw = row.get(col_map['country'], '')
                country = str(country_raw).strip() if pd.notna(country_raw) and str(country_raw).strip() else ''
            else:
                country = ''

            # Auto-detect country if not provided or empty
            if not country:
                country = self.detect_country_code(original_zip, city, street)
                country_detected = True

            # Normalize country to 2-letter code
            country_upper = country.upper()
            if country_upper in ('ITALY', 'ITALIA'):
                country = 'IT'
            elif country_upper in ('GERMANY', 'DEUTSCHLAND'):
                country = 'DE'
            elif country_upper in ('FRANCE'):
                country = 'FR'
            elif country_upper in ('SPAIN', 'ESPAÑA', 'ESPANA'):
                country = 'ES'
            elif country_upper in ('UNITED KINGDOM', 'UK', 'GREAT BRITAIN'):
                country = 'GB'
            elif len(country) == 2:
                country = country_upper
            else:
                # If still not a 2-letter code, try to detect
                country = self.detect_country_code(original_zip, city, street)
                country_detected = True

            # Skip non-IT countries (only validate Italian addresses)
            if country not in ('IT',):
                results.append(ValidationResult(
                    row_index=idx,
                    name=name,
                    city=city,
                    street=street,
                    original_zip=original_zip,
                    suggested_zip=original_zip,
                    confidence=100,
                    reason=f"Non-IT country ({country}) - skipped",
                    is_valid=True,
                    street_verified=True,
                    street_confidence=100,
                    country_code=country,
                    country_detected=country_detected,
                    phone_missing=phone_missing,
                    original_phone=original_phone,
                    cod_changed=cod_changed,
                    original_cod=original_cod
                ))
                skipped_count += 1
                continue

            # Validate full address (use combined street for better results)
            (
                is_valid, suggested_zip, confidence, reason,
                street_verified, suggested_street, street_confidence
            ) = self.validate_address(full_street, city, original_zip, country)

            # Cross-validate ZIP with province if state column is available
            if state and suggested_zip:
                province_valid, province_msg = self._validate_zip_province(suggested_zip, state)
                if not province_valid:
                    # ZIP doesn't match province - flag for review
                    is_valid = False
                    confidence = min(confidence, 70)
                    reason = f"{reason}. Warning: {province_msg}"

            # Determine ZIP action
            zip_auto_correct = not is_valid and confidence >= self.confidence_threshold and suggested_zip

            # Determine street action
            street_auto_correct = (
                not street_verified and
                suggested_street and
                street_confidence >= self.street_confidence_threshold
            )

            result = ValidationResult(
                row_index=idx,
                name=name,
                city=city,
                street=street,  # Keep original street (not combined) for output
                original_zip=original_zip,
                suggested_zip=suggested_zip,
                confidence=confidence,
                reason=reason,
                is_valid=is_valid,
                auto_corrected=zip_auto_correct,
                street_verified=street_verified,
                suggested_street=suggested_street,
                street_confidence=street_confidence,
                street_auto_corrected=street_auto_correct,
                country_code=country,
                country_detected=country_detected,
                phone_missing=phone_missing,
                original_phone=original_phone,
                cod_changed=cod_changed,
                original_cod=original_cod
            )
            results.append(result)

            # Count stats
            if is_valid:
                valid_count += 1
            elif zip_auto_correct:
                corrected_count += 1
            else:
                review_count += 1

            if street_verified:
                street_verified_count += 1
            elif street_auto_correct:
                street_corrected_count += 1

            # Rate limiting - only needed for Nominatim fallback
            if not self._photon_available:
                time.sleep(self.REQUEST_DELAY)

        logger.info(
            f"Validation complete: {valid_count} valid, {corrected_count} corrected, "
            f"{review_count} need review, {skipped_count} skipped"
        )
        logger.info(
            f"Streets: {street_verified_count} verified, {street_corrected_count} corrected"
        )

        report = ValidationReport(
            total_rows=total,
            valid_count=valid_count,
            corrected_count=corrected_count,
            review_count=review_count,
            skipped_count=skipped_count,
            results=results,
            street_verified_count=street_verified_count,
            street_corrected_count=street_corrected_count
        )
        return report, df

    def _validate_zip_province(self, zip_code: str, province: str) -> tuple[bool, str]:
        """
        Validate if ZIP code matches the Italian province.

        Args:
            zip_code: 5-digit Italian CAP
            province: 2-letter Italian province code (e.g., 'MI', 'RM')

        Returns:
            Tuple (is_valid, message)
        """
        if not province or not zip_code or len(zip_code) != 5:
            return True, ""  # Can't validate, assume OK

        province_upper = province.upper().strip()
        zip_prefix = zip_code[:2]

        if province_upper in self.ITALIAN_PROVINCE_ZIP:
            valid_prefixes = self.ITALIAN_PROVINCE_ZIP[province_upper]
            if zip_prefix in valid_prefixes:
                return True, f"ZIP matches province {province_upper}"
            else:
                return False, f"ZIP {zip_code} doesn't match province {province_upper} (expected {valid_prefixes[0]}xxx)"

        return True, ""  # Unknown province, assume OK

    # Default phone number to use when phone is missing
    DEFAULT_PHONE = "393445556667"

    def _map_columns(self, df: pd.DataFrame) -> dict:
        """Map DataFrame columns to expected fields."""
        col_map = {}
        columns_lower = {c.lower().strip(): c for c in df.columns}

        mappings = {
            'name': ['name', 'nome', 'customer name'],
            'street': ['street 1', 'street', 'address', 'indirizzo', 'via'],
            'street2': ['street 2', 'street2', 'address 2', 'indirizzo 2'],
            'city': ['city', 'città', 'citta'],
            'state': ['state', 'province', 'provincia', 'regione'],
            'zip': ['zip', 'cap', 'postal code', 'postcode', 'zip code'],
            'country': ['country', 'paese', 'nazione'],
            'phone': ['phone', 'telefono', 'tel', 'phone number', 'telephone'],
            'cash_on_delivery': ['cash on delivery', 'cod', 'contrassegno', 'cash_on_delivery'],
        }

        for field, possible_names in mappings.items():
            for name in possible_names:
                if name in columns_lower:
                    col_map[field] = columns_lower[name]
                    break

        return col_map

    def generate_corrected_excel(
        self,
        original_df: pd.DataFrame,
        report: ValidationReport
    ) -> bytes:
        """
        Generate corrected Excel with auto-corrections applied.
        Maintains the exact same column structure as the input file.

        Corrections applied:
        - ZIP: Corrected if auto_corrected flag is set
        - Street: Corrected if street_auto_corrected flag is set
        - Country: Filled from address detection if empty
        - Phone: Filled with default (393445556667) if empty
        - Cash on Delivery: Always set to 0

        Args:
            original_df: Original DataFrame
            report: Validation report

        Returns:
            Excel file as bytes
        """
        from openpyxl.utils import get_column_letter

        logger.info(f"Generating corrected Excel with {len(report.results)} rows")

        df = original_df.copy()
        col_map = self._map_columns(df)
        zip_col = col_map.get('zip')
        street_col = col_map.get('street')
        city_col = col_map.get('city')
        country_col = col_map.get('country')
        phone_col = col_map.get('phone')
        cod_col = col_map.get('cash_on_delivery')

        for result in report.results:
            # Correct ZIP
            if result.auto_corrected and result.suggested_zip and zip_col:
                df.at[result.row_index, zip_col] = result.suggested_zip

            # Correct street
            if result.street_auto_corrected and result.suggested_street and street_col:
                df.at[result.row_index, street_col] = result.suggested_street

            # Fill country code if column exists and value is empty
            if country_col:
                current_country = df.at[result.row_index, country_col]
                if not str(current_country).strip() or str(current_country).lower() == 'nan':
                    # Detect country from address
                    street_val = str(df.at[result.row_index, street_col]) if street_col else ''
                    city_val = str(df.at[result.row_index, city_col]) if city_col else ''
                    zip_val = str(df.at[result.row_index, zip_col]) if zip_col else ''
                    detected_country = self.detect_country_code(zip_val, city_val, street_val)
                    df.at[result.row_index, country_col] = detected_country

            # Fill phone with default if empty
            if phone_col:
                current_phone = df.at[result.row_index, phone_col]
                if pd.isna(current_phone) or not str(current_phone).strip() or str(current_phone).lower() == 'nan':
                    df.at[result.row_index, phone_col] = self.DEFAULT_PHONE

            # Set Cash on Delivery to 0 always
            if cod_col:
                df.at[result.row_index, cod_col] = 0

        # Ensure ZIP column is stored as string with leading zeros preserved
        if zip_col:
            df[zip_col] = df[zip_col].apply(
                lambda x: str(x).zfill(5) if pd.notna(x) and str(x).strip() else x
            )

        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Corrected')
            worksheet = writer.sheets['Corrected']

            # Format ZIP column as text to preserve leading zeros
            if zip_col:
                from openpyxl.styles import numbers
                zip_col_idx = list(df.columns).index(zip_col) + 1  # 1-indexed
                for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                    cell = worksheet.cell(row=row, column=zip_col_idx)
                    cell.number_format = numbers.FORMAT_TEXT

            # Auto-fit column widths
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                    len(str(col))
                ) + 2  # Add padding
                # Cap at 50 chars width
                column_width = min(max_length, 50)
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = column_width

        return output.getvalue()

    def generate_review_report(self, report: ValidationReport) -> bytes:
        """
        Generate Excel report for items needing manual review.

        Args:
            report: Validation report

        Returns:
            Excel file as bytes
        """
        from openpyxl.utils import get_column_letter

        review_items = []
        for r in report.results:
            if not r.is_valid or not r.street_verified:
                review_items.append({
                    'Row': r.row_index + 2,
                    'Name': r.name,
                    'City': r.city,
                    'Country': f"{r.country_code}{'*' if r.country_detected else ''}",
                    'Original Street': r.street,
                    'Suggested Street': r.suggested_street or '-',
                    'Street Conf.': f"{r.street_confidence}%" if r.street_confidence else '-',
                    'Street Action': (
                        'Verified' if r.street_verified else
                        'Auto-corrected' if r.street_auto_corrected else
                        'Review needed'
                    ),
                    'Original ZIP': r.original_zip,
                    'Suggested ZIP': r.suggested_zip or '-',
                    'ZIP Conf.': f"{r.confidence}%",
                    'ZIP Action': (
                        'Valid' if r.is_valid else
                        'Auto-corrected' if r.auto_corrected else
                        'Review needed'
                    ),
                    'Reason': r.reason
                })

        df = pd.DataFrame(review_items)
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Review')
            worksheet = writer.sheets['Review']

            # Auto-fit column widths
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                    len(col)
                ) + 2  # Add padding
                # Cap at 60 chars width
                column_width = min(max_length, 60)
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = column_width

        return output.getvalue()
